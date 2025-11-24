"""API Router for Materials."""
import logging
from datetime import datetime
from typing import List, Optional
import io

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook

from apps.database import get_db
from apps.config import settings
from apps.models.material import (
    MaterialCreate, MaterialUpdate, MaterialResponse,
    MaterialAddQuantity, MaterialPricingUpdate, MaterialHistoryEntry
)

logger = logging.getLogger('materials_router')
router = APIRouter(prefix="/api/materials", tags=["materials"])

VAT_MULTIPLIER = 1 + settings.VAT_RATE


def _material_row_to_response(row) -> dict:
    """Convert database row to response dict."""
    unit_cost = row['unit_cost_without_vat'] or 0
    total_cost = row['total_cost_without_vat'] or 0
    return {
        'id': row['id'],
        'name': row['name'],
        'category': row['category'],
        'unit': row['unit'],
        'quantity': row['quantity'] or 0,
        'is_active': bool(row['is_active']) if 'is_active' in row.keys() else True,
        'unit_cost_without_vat': unit_cost,
        'total_cost_without_vat': total_cost,
        'unit_cost_with_vat': round(unit_cost * VAT_MULTIPLIER, 2),
        'total_cost_with_vat': round(total_cost * VAT_MULTIPLIER, 2),
        'created_at': row['created_at'] if 'created_at' in row.keys() else None,
    }


async def _log_material_history(db, material_id: int, change_amount: float,
                                 change_type: str, performed_by: Optional[str] = None,
                                 description: Optional[str] = None):
    """Log a material history entry."""
    async with db.execute(
        "SELECT quantity FROM materials WHERE id = ?", (material_id,)
    ) as cursor:
        row = await cursor.fetchone()
        resulting_quantity = row['quantity'] if row else None

    await db.execute("""
        INSERT INTO material_history
        (material_id, change_type, change_amount, resulting_quantity, performed_by, description, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        material_id, change_type, change_amount, resulting_quantity,
        (performed_by or 'Неизвестно').strip() or 'Неизвестно',
        (description or '').strip(),
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ))


@router.get("", response_model=List[dict])
async def get_materials(active_only: bool = True):
    """Get all materials."""
    async with get_db() as db:
        query = """
            SELECT id, name, category, unit, quantity, is_active,
                   unit_cost_without_vat, total_cost_without_vat, created_at
            FROM materials
        """
        if active_only:
            query += " WHERE is_active = 1"
        query += " ORDER BY category, name"

        async with db.execute(query) as cursor:
            rows = await cursor.fetchall()
            return [_material_row_to_response(row) for row in rows]


@router.get("/history", response_model=List[MaterialHistoryEntry])
async def get_material_history(limit: int = 500):
    """Get material history."""
    async with get_db() as db:
        async with db.execute("""
            SELECT mh.id, mh.material_id, m.name, m.unit, mh.change_type,
                   mh.change_amount, mh.resulting_quantity, mh.performed_by,
                   mh.description, mh.created_at
            FROM material_history mh
            LEFT JOIN materials m ON m.id = mh.material_id
            ORDER BY mh.created_at DESC, mh.id DESC
            LIMIT ?
        """, (limit,)) as cursor:
            rows = await cursor.fetchall()
            return [
                MaterialHistoryEntry(
                    id=row['id'],
                    material_id=row['material_id'],
                    material_name=row['name'],
                    material_unit=row['unit'],
                    change_type=row['change_type'],
                    change_amount=row['change_amount'],
                    resulting_quantity=row['resulting_quantity'],
                    performed_by=row['performed_by'],
                    description=row['description'],
                    created_at=row['created_at']
                )
                for row in rows
            ]


@router.get("/export")
async def export_materials():
    """Export materials to Excel file."""
    async with get_db() as db:
        async with db.execute("""
            SELECT id, name, category, unit, quantity, is_active,
                   unit_cost_without_vat, total_cost_without_vat
            FROM materials ORDER BY category, name
        """) as cursor:
            rows = await cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Материалы"

    headers = ['ID', 'Название', 'Раздел', 'Единица', 'Количество',
               'Активен', 'Цена за ед. (без НДС)', 'Общая стоимость (без НДС)']
    ws.append(headers)

    for row in rows:
        ws.append([
            row['id'], row['name'], row['category'], row['unit'],
            row['quantity'] or 0, row['is_active'],
            row['unit_cost_without_vat'] or 0, row['total_cost_without_vat'] or 0
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=materials_export.xlsx"}
    )


@router.get("/template")
async def get_import_template():
    """Get Excel template for importing materials."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Материалы"

    headers = ['Название', 'Раздел', 'Единица', 'Количество', 'Цена за ед. (без НДС)']
    ws.append(headers)
    ws.append(['Пример материала', 'Категория', 'шт', 100, 150.50])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=materials_template.xlsx"}
    )


@router.post("/import")
async def import_materials(file: UploadFile = File(...)):
    """Import materials from Excel file."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "File must be Excel format (.xlsx or .xls)")

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    imported = 0
    errors = []

    async with get_db() as db:
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            try:
                name = str(row[0]).strip()
                category = str(row[1] if len(row) > 1 else '').strip()
                unit = str(row[2] if len(row) > 2 else 'шт').strip()
                quantity = float(row[3]) if len(row) > 3 and row[3] else 0
                unit_cost = float(row[4]) if len(row) > 4 and row[4] else 0
                total_cost = quantity * unit_cost

                # Ensure category exists
                if category:
                    await db.execute(
                        "INSERT OR IGNORE INTO categories (name, created_date) VALUES (?, ?)",
                        (category, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    )

                await db.execute("""
                    INSERT INTO materials (name, category, unit, quantity, is_active,
                                           unit_cost_without_vat, total_cost_without_vat, created_at)
                    VALUES (?, ?, ?, ?, 1, ?, ?, ?)
                """, (name, category, unit, quantity, unit_cost, total_cost,
                      datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                imported += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")

        await db.commit()

    return {"imported": imported, "errors": errors}


@router.get("/{material_id}", response_model=dict)
async def get_material(material_id: int):
    """Get a specific material by ID."""
    async with get_db() as db:
        async with db.execute("""
            SELECT id, name, category, unit, quantity, is_active,
                   unit_cost_without_vat, total_cost_without_vat, created_at
            FROM materials WHERE id = ?
        """, (material_id,)) as cursor:
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(404, "Material not found")
            return _material_row_to_response(row)


@router.post("", response_model=dict)
async def create_material(material: MaterialCreate):
    """Create a new material."""
    async with get_db() as db:
        # Ensure category exists
        if material.category:
            await db.execute(
                "INSERT OR IGNORE INTO categories (name, created_date) VALUES (?, ?)",
                (material.category, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )

        cursor = await db.execute("""
            INSERT INTO materials (name, category, unit, quantity, is_active,
                                   unit_cost_without_vat, total_cost_without_vat, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (material.name, material.category, material.unit, material.quantity,
              int(material.is_active), material.unit_cost_without_vat,
              material.total_cost_without_vat, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        await db.commit()

        material_id = cursor.lastrowid
        logger.info(f"Created material: {material.name} (ID: {material_id})")

        # Log history
        await _log_material_history(
            db, material_id, material.quantity,
            'Создание', 'Система', f'Создание материала {material.name}'
        )
        await db.commit()

        return await get_material(material_id)


@router.put("/{material_id}", response_model=dict)
async def update_material(material_id: int, material: MaterialUpdate):
    """Update an existing material."""
    async with get_db() as db:
        # Check if exists
        async with db.execute("SELECT id FROM materials WHERE id = ?", (material_id,)) as cursor:
            if not await cursor.fetchone():
                raise HTTPException(404, "Material not found")

        update_fields = []
        values = []

        if material.name is not None:
            update_fields.append("name = ?")
            values.append(material.name)
        if material.category is not None:
            update_fields.append("category = ?")
            values.append(material.category)
            await db.execute(
                "INSERT OR IGNORE INTO categories (name, created_date) VALUES (?, ?)",
                (material.category, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )
        if material.unit is not None:
            update_fields.append("unit = ?")
            values.append(material.unit)
        if material.quantity is not None:
            update_fields.append("quantity = ?")
            values.append(material.quantity)
        if material.is_active is not None:
            update_fields.append("is_active = ?")
            values.append(int(material.is_active))
        if material.unit_cost_without_vat is not None:
            update_fields.append("unit_cost_without_vat = ?")
            values.append(material.unit_cost_without_vat)
        if material.total_cost_without_vat is not None:
            update_fields.append("total_cost_without_vat = ?")
            values.append(material.total_cost_without_vat)

        if not update_fields:
            raise HTTPException(400, "No fields to update")

        values.append(material_id)
        query = f"UPDATE materials SET {', '.join(update_fields)} WHERE id = ?"

        await db.execute(query, values)
        await db.commit()
        logger.info(f"Updated material ID: {material_id}")
        return await get_material(material_id)


@router.put("/{material_id}/add-quantity", response_model=dict)
async def add_quantity(material_id: int, data: MaterialAddQuantity):
    """Add quantity to material stock."""
    async with get_db() as db:
        async with db.execute(
            "SELECT quantity FROM materials WHERE id = ?", (material_id,)
        ) as cursor:
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(404, "Material not found")

        new_quantity = (row['quantity'] or 0) + data.amount
        await db.execute(
            "UPDATE materials SET quantity = ? WHERE id = ?",
            (new_quantity, material_id)
        )

        await _log_material_history(
            db, material_id, data.amount, 'Приход',
            data.performed_by, data.description
        )
        await db.commit()

        logger.info(f"Added {data.amount} to material {material_id}. New quantity: {new_quantity}")
        return await get_material(material_id)


@router.get("/{material_id}/pricing", response_model=dict)
async def get_material_pricing(material_id: int):
    """Get material pricing info."""
    async with get_db() as db:
        async with db.execute("""
            SELECT id, name, unit_cost_without_vat, total_cost_without_vat, quantity
            FROM materials WHERE id = ?
        """, (material_id,)) as cursor:
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(404, "Material not found")

            unit_cost = row['unit_cost_without_vat'] or 0
            total_cost = row['total_cost_without_vat'] or 0
            return {
                'id': row['id'],
                'name': row['name'],
                'quantity': row['quantity'] or 0,
                'unit_cost_without_vat': unit_cost,
                'unit_cost_with_vat': round(unit_cost * VAT_MULTIPLIER, 2),
                'total_cost_without_vat': total_cost,
                'total_cost_with_vat': round(total_cost * VAT_MULTIPLIER, 2),
            }


@router.put("/{material_id}/pricing", response_model=dict)
async def update_material_pricing(material_id: int, data: MaterialPricingUpdate):
    """Update material pricing."""
    async with get_db() as db:
        async with db.execute(
            "SELECT quantity FROM materials WHERE id = ?", (material_id,)
        ) as cursor:
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(404, "Material not found")

        total_cost = data.unit_cost_without_vat * (row['quantity'] or 0)
        await db.execute("""
            UPDATE materials
            SET unit_cost_without_vat = ?, total_cost_without_vat = ?
            WHERE id = ?
        """, (data.unit_cost_without_vat, total_cost, material_id))
        await db.commit()

        logger.info(f"Updated pricing for material {material_id}")
        return await get_material_pricing(material_id)


@router.delete("/{material_id}")
async def delete_material(material_id: int):
    """Delete a material."""
    async with get_db() as db:
        # Delete from work_materials
        await db.execute("DELETE FROM work_materials WHERE material_id = ?", (material_id,))
        # Delete history
        await db.execute("DELETE FROM material_history WHERE material_id = ?", (material_id,))

        cursor = await db.execute("DELETE FROM materials WHERE id = ?", (material_id,))
        await db.commit()

        if cursor.rowcount == 0:
            raise HTTPException(404, "Material not found")

        logger.info(f"Deleted material ID: {material_id}")
        return {"message": "Material deleted successfully"}
