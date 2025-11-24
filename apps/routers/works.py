"""API Router for Works."""
import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook, load_workbook

from apps.database import get_db
from apps.config import settings
from apps.models.work import (
    WorkCreate, WorkUpdate, WorkResponse, WorkAddBalance,
    WorkMaterialsUpdate, WorkMaterialLink
)

logger = logging.getLogger('works_router')
router = APIRouter(prefix="/api/works", tags=["works"])

VAT_MULTIPLIER = 1 + settings.VAT_RATE


def _work_row_to_response(row) -> dict:
    """Convert database row to response dict."""
    unit_cost = row['unit_cost_without_vat'] or 0
    total_cost = row['total_cost_without_vat'] or 0
    return {
        'id': row['id'],
        'name': row['name'],
        'category': row['category'],
        'unit': row['unit'],
        'balance': row['balance'],
        'project_total': row['project_total'] or 0,
        'is_active': bool(row['is_active']),
        'unit_cost_without_vat': unit_cost,
        'total_cost_without_vat': total_cost,
        'unit_cost_with_vat': round(unit_cost * VAT_MULTIPLIER, 2),
        'total_cost_with_vat': round(total_cost * VAT_MULTIPLIER, 2),
    }


@router.get("", response_model=List[dict])
async def get_works(active_only: bool = True):
    """Get all works (optionally only active ones)."""
    async with get_db() as db:
        query = """
            SELECT id, name, category, unit, balance, project_total, is_active,
                   unit_cost_without_vat, total_cost_without_vat
            FROM works
        """
        if active_only:
            query += " WHERE is_active = 1"
        query += " ORDER BY category, name"

        async with db.execute(query) as cursor:
            rows = await cursor.fetchall()
            return [_work_row_to_response(row) for row in rows]


@router.get("/all", response_model=List[dict])
async def get_all_works():
    """Get all works including inactive."""
    return await get_works(active_only=False)


@router.get("/export")
async def export_works():
    """Export works to Excel file."""
    async with get_db() as db:
        async with db.execute("""
            SELECT id, name, category, unit, balance, project_total, is_active,
                   unit_cost_without_vat, total_cost_without_vat
            FROM works ORDER BY category, name
        """) as cursor:
            rows = await cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Работы"

    headers = ['ID', 'Название', 'Раздел', 'Единица', 'Баланс', 'Проект',
               'Активна', 'Цена за ед. (без НДС)', 'Общая стоимость (без НДС)']
    ws.append(headers)

    for row in rows:
        ws.append([
            row['id'], row['name'], row['category'], row['unit'],
            row['balance'], row['project_total'] or 0, row['is_active'],
            row['unit_cost_without_vat'] or 0, row['total_cost_without_vat'] or 0
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=works_export.xlsx"}
    )


@router.post("/import")
async def import_works(file: UploadFile = File(...)):
    """Import works from Excel file."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "File must be Excel format (.xlsx or .xls)")

    contents = await file.read()
    wb = load_workbook(io.BytesIO(contents))
    ws = wb.active

    imported = 0
    errors = []

    async with get_db() as db:
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            try:
                name = str(row[1] if len(row) > 1 else row[0]).strip()
                category = str(row[2] if len(row) > 2 else '').strip()
                unit = str(row[3] if len(row) > 3 else 'шт').strip()
                balance = float(row[4]) if len(row) > 4 and row[4] else 0
                project_total = float(row[5]) if len(row) > 5 and row[5] else 0
                is_active = int(row[6]) if len(row) > 6 and row[6] is not None else 1
                unit_cost = float(row[7]) if len(row) > 7 and row[7] else 0
                total_cost = float(row[8]) if len(row) > 8 and row[8] else 0

                # Ensure category exists
                if category:
                    await db.execute(
                        "INSERT OR IGNORE INTO categories (name, created_date) VALUES (?, ?)",
                        (category, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                    )

                await db.execute("""
                    INSERT INTO works (name, category, unit, balance, project_total, is_active,
                                       unit_cost_without_vat, total_cost_without_vat)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(name) DO UPDATE SET
                        category = excluded.category,
                        unit = excluded.unit,
                        balance = excluded.balance,
                        project_total = excluded.project_total,
                        is_active = excluded.is_active,
                        unit_cost_without_vat = excluded.unit_cost_without_vat,
                        total_cost_without_vat = excluded.total_cost_without_vat
                """, (name, category, unit, balance, project_total, is_active, unit_cost, total_cost))
                imported += 1
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")

        await db.commit()

    return {"imported": imported, "errors": errors}


@router.get("/{work_id}", response_model=dict)
async def get_work(work_id: int):
    """Get a specific work by ID."""
    async with get_db() as db:
        async with db.execute("""
            SELECT id, name, category, unit, balance, project_total, is_active,
                   unit_cost_without_vat, total_cost_without_vat
            FROM works WHERE id = ?
        """, (work_id,)) as cursor:
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(404, "Work not found")
            return _work_row_to_response(row)


@router.post("", response_model=dict)
async def create_work(work: WorkCreate):
    """Create a new work."""
    async with get_db() as db:
        # Ensure category exists
        if work.category:
            await db.execute(
                "INSERT OR IGNORE INTO categories (name, created_date) VALUES (?, ?)",
                (work.category, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )

        try:
            cursor = await db.execute("""
                INSERT INTO works (name, category, unit, balance, project_total, is_active,
                                   unit_cost_without_vat, total_cost_without_vat)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (work.name, work.category, work.unit, work.balance, work.project_total,
                  int(work.is_active), work.unit_cost_without_vat, work.total_cost_without_vat))
            await db.commit()
            work_id = cursor.lastrowid
            logger.info(f"Created work: {work.name} (ID: {work_id})")
            return await get_work(work_id)
        except Exception as e:
            if "UNIQUE constraint failed" in str(e):
                raise HTTPException(400, "Work with this name already exists")
            raise HTTPException(500, str(e))


@router.put("/{work_id}", response_model=dict)
async def update_work(work_id: int, work: WorkUpdate):
    """Update an existing work."""
    async with get_db() as db:
        # Check if work exists
        async with db.execute("SELECT id FROM works WHERE id = ?", (work_id,)) as cursor:
            if not await cursor.fetchone():
                raise HTTPException(404, "Work not found")

        update_fields = []
        values = []

        if work.name is not None:
            update_fields.append("name = ?")
            values.append(work.name)
        if work.category is not None:
            update_fields.append("category = ?")
            values.append(work.category)
            # Ensure category exists
            await db.execute(
                "INSERT OR IGNORE INTO categories (name, created_date) VALUES (?, ?)",
                (work.category, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )
        if work.unit is not None:
            update_fields.append("unit = ?")
            values.append(work.unit)
        if work.balance is not None:
            update_fields.append("balance = ?")
            values.append(work.balance)
        if work.project_total is not None:
            update_fields.append("project_total = ?")
            values.append(work.project_total)
        if work.is_active is not None:
            update_fields.append("is_active = ?")
            values.append(int(work.is_active))
        if work.unit_cost_without_vat is not None:
            update_fields.append("unit_cost_without_vat = ?")
            values.append(work.unit_cost_without_vat)
        if work.total_cost_without_vat is not None:
            update_fields.append("total_cost_without_vat = ?")
            values.append(work.total_cost_without_vat)

        if not update_fields:
            raise HTTPException(400, "No fields to update")

        values.append(work_id)
        query = f"UPDATE works SET {', '.join(update_fields)} WHERE id = ?"

        try:
            await db.execute(query, values)
            await db.commit()
            logger.info(f"Updated work ID: {work_id}")
            return await get_work(work_id)
        except Exception as e:
            if "UNIQUE constraint failed" in str(e):
                raise HTTPException(400, "Work with this name already exists")
            raise HTTPException(500, str(e))


@router.put("/{work_id}/add-balance", response_model=dict)
async def add_balance(work_id: int, data: WorkAddBalance):
    """Add balance to a work."""
    async with get_db() as db:
        async with db.execute("SELECT balance FROM works WHERE id = ?", (work_id,)) as cursor:
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(404, "Work not found")

        new_balance = (row['balance'] or 0) + data.amount
        await db.execute("UPDATE works SET balance = ? WHERE id = ?", (new_balance, work_id))
        await db.commit()
        logger.info(f"Added {data.amount} to work {work_id} balance. New balance: {new_balance}")
        return await get_work(work_id)


@router.delete("/{work_id}")
async def delete_work(work_id: int):
    """Delete a work."""
    async with get_db() as db:
        # Delete related work_materials first
        await db.execute("DELETE FROM work_materials WHERE work_id = ?", (work_id,))

        cursor = await db.execute("DELETE FROM works WHERE id = ?", (work_id,))
        await db.commit()

        if cursor.rowcount == 0:
            raise HTTPException(404, "Work not found")

        logger.info(f"Deleted work ID: {work_id}")
        return {"message": "Work deleted successfully"}


@router.get("/{work_id}/materials", response_model=List[dict])
async def get_work_materials(work_id: int):
    """Get materials linked to a work."""
    async with get_db() as db:
        # Check if work exists
        async with db.execute("SELECT id FROM works WHERE id = ?", (work_id,)) as cursor:
            if not await cursor.fetchone():
                raise HTTPException(404, "Work not found")

        async with db.execute("""
            SELECT wm.material_id, wm.quantity_per_unit, m.name, m.unit, m.quantity
            FROM work_materials wm
            JOIN materials m ON wm.material_id = m.id
            WHERE wm.work_id = ?
        """, (work_id,)) as cursor:
            rows = await cursor.fetchall()
            return [{
                'material_id': row['material_id'],
                'quantity_per_unit': row['quantity_per_unit'],
                'material_name': row['name'],
                'material_unit': row['unit'],
                'available_quantity': row['quantity']
            } for row in rows]


@router.put("/{work_id}/materials")
async def update_work_materials(work_id: int, data: WorkMaterialsUpdate):
    """Update materials linked to a work."""
    async with get_db() as db:
        # Check if work exists
        async with db.execute("SELECT id FROM works WHERE id = ?", (work_id,)) as cursor:
            if not await cursor.fetchone():
                raise HTTPException(404, "Work not found")

        # Replace all materials
        await db.execute("DELETE FROM work_materials WHERE work_id = ?", (work_id,))

        for mat in data.materials:
            await db.execute("""
                INSERT INTO work_materials (work_id, material_id, quantity_per_unit)
                VALUES (?, ?, ?)
            """, (work_id, mat.material_id, mat.quantity_per_unit))

        await db.commit()
        logger.info(f"Updated materials for work {work_id}")
        return await get_work_materials(work_id)
