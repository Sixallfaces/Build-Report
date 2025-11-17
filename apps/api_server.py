# api_server.py
import aiosqlite
from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from typing import List, Optional
import json
import logging
import os
import re
from datetime import datetime
import traceback
import hashlib
import secrets
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
import io
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import requests

# --- Настройки ---
DB_PATH = '/opt/stroykontrol/database/stroykontrol.db'
API_HOST = '127.0.0.1'
API_PORT = 8080

# --- Настройки Яндекс.Диска ---
YANDEX_DISK_TOKEN = os.getenv('YANDEX_DISK_TOKEN')
YANDEX_DISK_BASE_FOLDER = os.getenv('YANDEX_DISK_BASE_FOLDER', 'StroyKontrol')


# --- Настройка логирования ---
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger('api_server')

# --- Приложение FastAPI ---
app = FastAPI(title="StroyKontrol API", version="1.0.0")

# --- CORS Middleware ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["https://build-report.ru"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Вспомогательные функции для работы с Яндекс.Диском ---
def _get_yandex_headers():
    if not YANDEX_DISK_TOKEN:
        logger.error("❌ Токен Яндекс.Диска не задан. Укажите переменную окружения YANDEX_DISK_TOKEN")
        return None
    return {
        'Authorization': f'OAuth {YANDEX_DISK_TOKEN}',
        'Content-Type': 'application/json'
    }


def sanitize_folder_component(component: str) -> str:
    safe = re.sub(r'[^\w\-]', '_', component or '')
    return safe or 'unknown'


def setup_yandex_disk() -> bool:
    headers = _get_yandex_headers()
    if headers is None:
        return False
    try:
        response = requests.get('https://cloud-api.yandex.net/v1/disk/', headers=headers, timeout=10)
        if response.status_code == 200:
            return True
        logger.error(f"❌ Ошибка подключения к Яндекс.Диску: {response.status_code} - {response.text}")
    except requests.RequestException as exc:
        logger.error(f"❌ Исключение при подключении к Яндекс.Диску: {exc}")
    return False


def create_yandex_folder(folder_path: str) -> bool:
    headers = _get_yandex_headers()
    if headers is None:
        return False

    if not folder_path.startswith('/'):
        folder_path = '/' + folder_path

    params = {'path': folder_path}
    try:
        response = requests.put(
            'https://cloud-api.yandex.net/v1/disk/resources',
            headers=headers,
            params=params,
            timeout=10
        )
        if response.status_code in (200, 201):
            logger.info(f"✅ Папка на Яндекс.Диске создана: {folder_path}")
            return True
        if response.status_code == 409:
            logger.info(f"ℹ️ Папка на Яндекс.Диске уже существует: {folder_path}")
            return True
        logger.error(f"❌ Ошибка создания папки на Яндекс.Диске: {response.status_code} - {response.text}")
    except requests.RequestException as exc:
        logger.error(f"❌ Исключение при создании папки на Яндекс.Диске: {exc}")
    return False


def publish_yandex_folder(folder_path: str) -> Optional[str]:
    headers = _get_yandex_headers()
    if headers is None:
        return None

    if folder_path.startswith('/'):
        folder_path = folder_path[1:]

    try:
        publish_response = requests.put(
            'https://cloud-api.yandex.net/v1/disk/resources/publish',
            headers=headers,
            params={'path': folder_path},
            timeout=10
        )
        if publish_response.status_code not in (200, 201):
            logger.error(
                f"❌ Не удалось опубликовать папку {folder_path}: {publish_response.status_code} - {publish_response.text}"
            )
            return None

        info_response = requests.get(
            'https://cloud-api.yandex.net/v1/disk/resources',
            headers=headers,
            params={'path': folder_path, 'fields': 'public_url'},
            timeout=10
        )
        if info_response.status_code == 200:
            public_url = info_response.json().get('public_url')
            if public_url:
                logger.info(f"🔗 Получена ссылка на папку: {public_url}")
                return public_url
        logger.error(
            f"❌ Не удалось получить ссылку на папку {folder_path}: {info_response.status_code} - {info_response.text}"
        )
    except requests.RequestException as exc:
        logger.error(f"❌ Исключение при публикации папки на Яндекс.Диске: {exc}")
    return None


async def ensure_report_folder(db, foreman_id: Optional[int], report_date: str) -> Optional[str]:
    if foreman_id is None or report_date is None:
        return None

    if not setup_yandex_disk():
        return None

    async with db.execute(
        "SELECT first_name FROM foremen WHERE id = ?",
        (foreman_id,)
    ) as cursor:
        row = await cursor.fetchone()

    if not row:
        logger.warning(f"⚠️ Не удалось найти бригадира ID {foreman_id} для создания папки отчета")
        return None

    foreman_name = sanitize_folder_component(row[0])
    base_folder = sanitize_folder_component(YANDEX_DISK_BASE_FOLDER or 'StroyKontrol')
    date_folder = sanitize_folder_component(report_date)

    base_folder_path = f"/{base_folder}"
    if not create_yandex_folder(base_folder_path):
        return None

    date_folder_path = f"{base_folder_path}/{date_folder}"
    if not create_yandex_folder(date_folder_path):
        return None

    foreman_folder_path = f"{date_folder_path}/{foreman_name}_ID_{foreman_id}"
    if not create_yandex_folder(foreman_folder_path):
        return None

    return publish_yandex_folder(foreman_folder_path)

async def ensure_work_reports_verification_column():
    """Гарантирует наличие колонки is_verified в таблице work_reports."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute("PRAGMA table_info(work_reports)") as cursor:
                columns = [row[1] for row in await cursor.fetchall()]

            if 'is_verified' not in columns:
                await db.execute(
                    "ALTER TABLE work_reports ADD COLUMN is_verified INTEGER NOT NULL DEFAULT 0"
                )
                await db.commit()
                logger.info("✅ Добавлена колонка is_verified в таблицу work_reports")
    except Exception as exc:
        logger.error(f"⚠️ Ошибка при добавлении колонки is_verified: {exc}")

async def ensure_work_pricing_columns():
    """Гарантирует наличие колонок цен в таблице works."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute("PRAGMA table_info(works)") as cursor:
                columns = [row[1] for row in await cursor.fetchall()]

            if 'unit_cost_without_vat' not in columns:
                await db.execute(
                    "ALTER TABLE works ADD COLUMN unit_cost_without_vat REAL NOT NULL DEFAULT 0"
                )
                logger.info("✅ Добавлена колонка unit_cost_without_vat в таблицу works")

            if 'total_cost_without_vat' not in columns:
                await db.execute(
                    "ALTER TABLE works ADD COLUMN total_cost_without_vat REAL NOT NULL DEFAULT 0"
                )
                logger.info("✅ Добавлена колонка total_cost_without_vat в таблицу works")

            await db.commit()
    except Exception as exc:
        logger.error(f"⚠️ Ошибка при добавлении колонок цен в works: {exc}")


async def ensure_material_pricing_columns():
    """Гарантирует наличие колонок цен в таблице materials."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute("PRAGMA table_info(materials)") as cursor:
                columns = [row[1] for row in await cursor.fetchall()]

            if 'unit_cost_without_vat' not in columns:
                await db.execute(
                    "ALTER TABLE materials ADD COLUMN unit_cost_without_vat REAL NOT NULL DEFAULT 0"
                )
                logger.info("✅ Добавлена колонка unit_cost_without_vat в таблицу materials")

            if 'total_cost_without_vat' not in columns:
                await db.execute(
                    "ALTER TABLE materials ADD COLUMN total_cost_without_vat REAL NOT NULL DEFAULT 0"
                )
                logger.info("✅ Добавлена колонка total_cost_without_vat в таблицу materials")

            await db.commit()
    except Exception as exc:
        logger.error(f"⚠️ Ошибка при добавлении колонок цен в materials: {exc}")

# Хэш-функция для паролей
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

# Таблица пользователей сайта
async def init_site_users_table():
    """Создает таблицу для пользователей сайта"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute('''
            CREATE TABLE IF NOT EXISTS site_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                is_active INTEGER NOT NULL DEFAULT 1,
                created_date TEXT NOT NULL,
                last_login TEXT
            )
        ''')
        await db.commit()

# ========== ФУНКЦИИ ДЛЯ РАБОТ ==========
async def get_active_works_from_db():
    """Получает список активных работ из базы данных - ДЛЯ БОТА."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT id, name, category, unit, balance, project_total, is_active, "
                "unit_cost_without_vat, total_cost_without_vat FROM works WHERE is_active = 1"
            ) as cursor:
                rows = await cursor.fetchall()
                works = []
                for row in rows:
                    (
                        work_id,
                        name,
                        category,
                        unit,
                        balance,
                        project_total,
                        is_active,
                        unit_cost_without_vat,
                        total_cost_without_vat,
                    ) = row
                    works.append({
                        'id': work_id,
                        'Название работы': name,
                        'Раздел': category,
                        'Единица измерения': unit,
                        'На балансе': balance,
                        'Проект': project_total,  # НОВОЕ ПОЛЕ
                        'is_active': bool(is_active),
                        'Стоимость за единицу': unit_cost_without_vat or 0,
                        'Всего стоимость': total_cost_without_vat or 0,
                    })
                logger.info(f"🔍 Найдено активных работ в БД: {len(works)}")
                return works
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения активных работ: {e}")
        return []

async def get_all_works_from_db():
    """Получает список ВСЕХ работ из базы данных - ДЛЯ САЙТА."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT id, name, category, unit, balance, project_total, is_active, "
                "unit_cost_without_vat, total_cost_without_vat FROM works"
            ) as cursor:
                rows = await cursor.fetchall()
                works = []
                for row in rows:
                    (
                        work_id,
                        name,
                        category,
                        unit,
                        balance,
                        project_total,
                        is_active,
                        unit_cost_without_vat,
                        total_cost_without_vat,
                    ) = row
                    works.append({
                        'id': work_id,
                        'Название работы': name,
                        'Раздел': category,
                        'Единица измерения': unit,
                        'На балансе': balance,
                        'Проект': project_total,  # НОВОЕ ПОЛЕ
                        'is_active': bool(is_active),
                        'Стоимость за единицу': unit_cost_without_vat or 0,
                        'Всего стоимость': total_cost_without_vat or 0,
                    })
                logger.info(f"🔍 Найдено всех работ в БД: {len(works)}")
                return works
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения всех работ: {e}")
        return []

async def get_work_by_id(work_id: int):
    """Получает конкретную работу по ID."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT id, name, category, unit, balance, project_total, is_active, "
                "unit_cost_without_vat, total_cost_without_vat FROM works WHERE id = ?",
                (work_id,)
            ) as cursor:
                row = await cursor.fetchone()
                if row:
                    (
                        work_id,
                        name,
                        category,
                        unit,
                        balance,
                        project_total,
                        is_active,
                        unit_cost_without_vat,
                        total_cost_without_vat,
                    ) = row
                    return {
                        'id': work_id,
                        'Название работы': name,
                        'Раздел': category,
                        'Единица измерения': unit,
                        'На балансе': balance,
                        'Проект': project_total,  # НОВОЕ ПОЛЕ
                        'is_active': bool(is_active),
                        'Стоимость за единицу (без НДС)': unit_cost_without_vat or 0,
                        'Стоимость за единицу (с НДС)': round((unit_cost_without_vat or 0) * VAT_MULTIPLIER, 2),
                        'Всего стоимость (без НДС)': total_cost_without_vat or 0,
                        'Всего стоимость (с НДС)': round((total_cost_without_vat or 0) * VAT_MULTIPLIER, 2),
                    }
        return None
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения работы по ID {work_id}: {e}")
        return None

async def insert_work_to_db(work_data: dict):
    """Добавляет новую работу в базу данных."""
    try:
        logger.info(f"DEBUG: insert_work_to_db пытается вставить: {work_data}")
        await ensure_category_exists_in_db(work_data.get('category'))
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "INSERT INTO works (name, category, unit, balance, project_total, is_active, "
                "unit_cost_without_vat, total_cost_without_vat) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    work_data['name'],
                    work_data['category'],
                    work_data['unit'],
                    work_data['balance'],
                    work_data.get('project_total', 0),
                    work_data['is_active'],
                    work_data.get('unit_cost_without_vat', 0),
                    work_data.get('total_cost_without_vat', 0),
                )
            )
            await db.commit()
            work_id = db.last_insert_rowid()
            logger.info(f"🏗️ Добавлена новая работа: {work_data['name']} (ID: {work_id})")
            return work_id
    except aiosqlite.IntegrityError as e:
        logger.error(f"❌ Ошибка целостности базы данных при вставке {work_data}: {e}")
        raise
    except Exception as e:
        logger.error(f"⚠️ Неожиданная ошибка добавления работы {work_data}: {e}")
        logger.error(traceback.format_exc())
        return None

async def update_work_in_db(work_id: int, work_data: dict):
    """Обновляет существующую работу в базе данных."""
    try:
        await ensure_category_exists_in_db(work_data.get('category'))
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "UPDATE works SET name = ?, category = ?, unit = ?, balance = ?, project_total = ?, "
                "is_active = ?, unit_cost_without_vat = ?, total_cost_without_vat = ? WHERE id = ?",
                (
                    work_data['name'],
                    work_data['category'],
                    work_data['unit'],
                    work_data['balance'],
                    work_data.get('project_total', 0),
                    work_data['is_active'],
                    work_data.get('unit_cost_without_vat', 0),
                    work_data.get('total_cost_without_vat', 0),
                    work_id,
                )
            )
            await db.commit()
            if db.rowcount > 0:
                logger.info(f"🏗️ Обновлена работа ID: {work_id}")
                return True
        return False
    except Exception as e:
        logger.error(f"⚠️ Ошибка обновления работы ID {work_id}: {e}")
        return False
    
async def add_balance_to_work_in_db(work_id: int, amount: float):
    """Увеличивает баланс работы на указанную величину."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")

                async with db.execute(
                    "SELECT balance FROM works WHERE id = ?",
                    (work_id,)
                ) as cursor:
                    row = await cursor.fetchone()
                    if not row:
                        await db.rollback()
                        return None

                    current_balance = row[0] if row[0] is not None else 0

                new_balance = current_balance + amount

                await db.execute(
                    "UPDATE works SET balance = ? WHERE id = ?",
                    (new_balance, work_id)
                )
                await db.commit()
                logger.info(
                    f"🏗️ Баланс работы ID: {work_id} увеличен на {amount}. Новый баланс: {new_balance}"
                )
                return new_balance
            except Exception as inner_error:
                await db.rollback()
                logger.error(
                    f"⚠️ Ошибка при увеличении баланса работы ID {work_id}: {inner_error}"
                )
                return None
    except Exception as e:
        logger.error(f"⚠️ Ошибка соединения при увеличении баланса работы ID {work_id}: {e}")
        return None 

async def delete_work_from_db(work_id: int):
    """Удаляет работу из базы данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")
                await db.execute("DELETE FROM work_materials WHERE work_id = ?", (work_id,))
                await db.execute("DELETE FROM works WHERE id = ?", (work_id,))
                await db.commit()
                if db.total_changes > 0:
                    logger.info(f"🗑️ Удалена работа ID: {work_id}")
                    return True
            except Exception as inner_error:
                await db.rollback()
                logger.error(f"⚠️ Ошибка при удалении работы ID {work_id}: {inner_error}")
                return False
        return False
    except Exception as e:
        logger.error(f"⚠️ Ошибка удаления работы ID {work_id}: {e}")
        return False

# ========== ФУНКЦИИ ДЛЯ БРИГАДИРОВ ==========
async def get_foremen_from_db():
    """Получает список всех бригадиров из базы данных (и активных, и неактивных)."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT id, first_name, last_name, username, registration_date, is_active FROM foremen"
            ) as cursor:
                rows = await cursor.fetchall()
                foremen = []
                for row in rows:
                    foreman_id, full_name, position, username, reg_date, is_active = row
                    foremen.append({
                        'id': foreman_id,
                        'full_name': full_name,
                        'position': position or '',
                        'first_name': full_name,  # для обратной совместимости
                        'last_name': position or '',
                        'username': username,
                        'registration_date': reg_date,
                        'is_active': bool(is_active) if is_active is not None else True
                    })
                logger.info(f"👥 Найдено бригадиров в БД: {len(foremen)}")
                return foremen
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения бригадиров: {e}")
        return []
    
async def get_foreman_display_name(db, foreman_id: Optional[int]) -> str:
    """Возвращает отображаемое имя бригадира"""
    if foreman_id is None:
        return 'Неизвестный бригадир'

    async with db.execute(
        "SELECT first_name, last_name FROM foremen WHERE id = ?",
        (foreman_id,)
    ) as cursor:
        row = await cursor.fetchone()
        if row:
            first_name, last_name = row
            parts = [part for part in [first_name, last_name] if part]
            if parts:
                return f"Бригадир {' '.join(parts)}"
    return f"Бригадир ID {foreman_id}"

async def create_foreman_in_db(foreman_data: dict):
    """Создает нового бригадира в базе данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "INSERT INTO foremen (first_name, last_name, username, registration_date, is_active) VALUES (?, ?, ?, ?, ?)",
                (foreman_data['full_name'], foreman_data['position'],
                 foreman_data.get('username', ''), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 1)  # is_active = 1 по умолчанию
            )
            await db.commit()
            foreman_id = db.last_insert_rowid()
            logger.info(f"👤 Добавлен новый бригадир: {foreman_data['full_name']} ({foreman_data['position']}) (ID: {foreman_id})")
            return foreman_id
    except Exception as e:
        logger.error(f"❌ Ошибка добавления бригадира: {e}")
        return None

# Обновим функцию update_foreman_in_db для поддержки is_active
async def update_foreman_in_db(foreman_id: int, foreman_data: dict):
    """Обновляет данные бригадира в базе данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT first_name, last_name, username, is_active FROM foremen WHERE id = ?",
                (foreman_id,)
            ) as cursor:
                existing = await cursor.fetchone()

            if not existing:
                return False

            existing_first, existing_last, existing_username, existing_is_active = existing

            first_name = (
                foreman_data.get('full_name')
                or foreman_data.get('first_name')
                or existing_first
            )
            last_name = (
                foreman_data.get('position')
                or foreman_data.get('last_name')
                or existing_last
            )
            username = existing_username or ''
            is_active = foreman_data.get('is_active', existing_is_active)

            await db.execute(
                "UPDATE foremen SET first_name = ?, last_name = ?, username = ?, is_active = ? WHERE id = ?",
                (first_name, last_name, username, is_active, foreman_id)

            )
            await db.commit()
            logger.info(
                f"👤 Обновлен бригадир ID: {foreman_id}, is_active: {is_active}"
            )
            return True
        return False
    except Exception as e:
        logger.error(f"❌ Ошибка обновления бригадира ID {foreman_id}: {e}")
        return False

async def delete_foreman_from_db(foreman_id: int):
    """Удаляет бригадира из базы данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            # Проверяем, есть ли отчеты у бригадира
            async with db.execute(
                "SELECT COUNT(*) FROM work_reports WHERE foreman_id = ?", 
                (foreman_id,)
            ) as cursor:
                report_count = await cursor.fetchone()
                if report_count and report_count[0] > 0:
                    return False, "Нельзя удалить бригадира, у которого есть отчеты"
            
            cursor = await db.execute("DELETE FROM foremen WHERE id = ?", (foreman_id,))
            await db.commit()
            if cursor.rowcount and cursor.rowcount > 0:
                logger.info(f"🗑️ Удален бригадир ID: {foreman_id}")
                return True, "Бригадир успешно удален"
        return False, "Бригадир не найден"
    except Exception as e:
        logger.error(f"❌ Ошибка удаления бригадира ID {foreman_id}: {e}")
        return False, f"Ошибка удаления: {str(e)}"

async def ensure_foreman_sections_table():
    """Создает таблицу связей бригадир-раздел при необходимости."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                """
                CREATE TABLE IF NOT EXISTS foreman_sections (
                    foreman_id INTEGER NOT NULL,
                    category_id INTEGER NOT NULL,
                    PRIMARY KEY (foreman_id, category_id),
                    FOREIGN KEY (foreman_id) REFERENCES foremen(id) ON DELETE CASCADE,
                    FOREIGN KEY (category_id) REFERENCES categories(id) ON DELETE CASCADE
                )
                """
            )
            await db.commit()
    except Exception as exc:
        logger.error(f"⚠️ Не удалось гарантировать наличие таблицы foreman_sections: {exc}")


async def fetch_foreman_sections(db, foreman_id: int) -> List[dict]:
    """Возвращает список разделов для бригадира, используя существующее соединение."""
    async with db.execute(
        """
        SELECT c.id, c.name
        FROM foreman_sections fs
        JOIN categories c ON fs.category_id = c.id
        WHERE fs.foreman_id = ?
        ORDER BY c.name
        """,
        (foreman_id,)
    ) as cursor:
        rows = await cursor.fetchall()
        return [{"id": row[0], "name": row[1]} for row in rows]


async def get_foreman_sections_from_db(foreman_id: int) -> List[dict]:
    """Получает список разделов, закрепленных за бригадиром."""
    await ensure_foreman_sections_table()
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            return await fetch_foreman_sections(db, foreman_id)
    except Exception as exc:
        logger.error(f"⚠️ Ошибка получения разделов для бригадира {foreman_id}: {exc}")
        return []


async def replace_foreman_sections_for_foreman(foreman_id: int, category_ids: List[int]):
    """Полностью заменяет список разделов, закрепленных за бригадиром."""
    await ensure_foreman_sections_table()

    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT 1 FROM foremen WHERE id = ?",
                (foreman_id,)
            ) as cursor:
                if await cursor.fetchone() is None:
                    return False, "Бригадир не найден", None

            unique_ids = []
            seen = set()
            for raw_id in category_ids:
                try:
                    category_id = int(raw_id)
                except (TypeError, ValueError):
                    return False, "Идентификатор раздела должен быть числом", None
                if category_id <= 0:
                    return False, "Идентификатор раздела должен быть положительным", None
                if category_id not in seen:
                    seen.add(category_id)
                    unique_ids.append(category_id)

            if unique_ids:
                placeholders = ",".join(["?"] * len(unique_ids))
                async with db.execute(
                    f"SELECT id FROM categories WHERE id IN ({placeholders})",
                    unique_ids
                ) as cursor:
                    existing_ids = {row[0] for row in await cursor.fetchall()}
                missing = [category_id for category_id in unique_ids if category_id not in existing_ids]
                if missing:
                    return False, f"Некоторые разделы не найдены: {', '.join(map(str, missing))}", None

            try:
                await db.execute("BEGIN")
                await db.execute("DELETE FROM foreman_sections WHERE foreman_id = ?", (foreman_id,))
                for category_id in unique_ids:
                    await db.execute(
                        "INSERT INTO foreman_sections (foreman_id, category_id) VALUES (?, ?)",
                        (foreman_id, category_id)
                    )
                await db.commit()
                logger.info(f"🔗 Обновлены разделы для бригадира ID: {foreman_id}")
            except Exception as exc:
                await db.rollback()
                logger.error(f"⚠️ Ошибка транзакции обновления разделов для бригадира {foreman_id}: {exc}")
                return False, "Не удалось обновить разделы бригадира", None

            updated_sections = await fetch_foreman_sections(db, foreman_id)
            return True, None, updated_sections
    except Exception as exc:
        logger.error(f"⚠️ Ошибка обновления разделов бригадира {foreman_id}: {exc}")
        return False, str(exc), None


async def foreman_exists(foreman_id: int) -> bool:
    """Проверяет наличие бригадира в базе данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute("SELECT 1 FROM foremen WHERE id = ?", (foreman_id,)) as cursor:
                return await cursor.fetchone() is not None
    except Exception as exc:
        logger.error(f"⚠️ Ошибка проверки существования бригадира {foreman_id}: {exc}")
        return False

# ========== ФУНКЦИИ ДЛЯ РАЗДЕЛОВ ==========
async def init_categories_table():
    """Создает таблицу для разделов"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                created_date TEXT NOT NULL
            )
        ''')
        await db.commit()

async def get_categories_from_db():
    """Получает список всех разделов из базы данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT id, name, created_date FROM categories ORDER BY name"
            ) as cursor:
                rows = await cursor.fetchall()
                categories = []
                for row in rows:
                    category_id, name, created_date = row
                    categories.append({
                        'id': category_id,
                        'name': name,
                        'created_date': created_date
                    })
                logger.info(f"📂 Найдено разделов в БД: {len(categories)}")
                return categories
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения разделов: {e}")
        return []

async def create_category_in_db(category_data: dict):
    """Добавляет новый раздел в базу данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "INSERT INTO categories (name, created_date) VALUES (?, ?)",
                (category_data['name'], datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            )
            await db.commit()
            category_id = db.last_insert_rowid()
            logger.info(f"📂 Добавлен новый раздел: {category_data['name']} (ID: {category_id})")
            return category_id
    except aiosqlite.IntegrityError:
        raise HTTPException(status_code=400, detail="Раздел с таким названием уже существует")
    except Exception as e:
        logger.error(f"⚠️ Ошибка добавления раздела: {e}")
        return None
    
async def update_category_in_db(category_id: int, new_name: str):
    """Обновляет название раздела и связанные записи."""
    normalized_name = (new_name or '').strip()
    if not normalized_name:
        return False, "Отсутствует название раздела"

    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT name FROM categories WHERE id = ?",
                (category_id,),
            ) as cursor:
                row = await cursor.fetchone()
                if not row:
                    return False, "Раздел не найден"
                current_name = row[0]

            if current_name == normalized_name:
                return True, "Название раздела не изменилось"

            try:
                await db.execute(
                    "UPDATE categories SET name = ? WHERE id = ?",
                    (normalized_name, category_id),
                )
            except aiosqlite.IntegrityError:
                return False, "Раздел с таким названием уже существует"

            await db.execute(
                "UPDATE works SET category = ? WHERE category = ?",
                (normalized_name, current_name),
            )
            await db.execute(
                "UPDATE materials SET category = ? WHERE category = ?",
                (normalized_name, current_name),
            )
            await db.commit()

            logger.info(
                "✏️ Обновлен раздел ID %s: '%s' → '%s'",
                category_id,
                current_name,
                normalized_name,
            )
            return True, "Раздел успешно обновлен"
    except Exception as exc:
        logger.error(f"⚠️ Ошибка обновления раздела ID {category_id}: {exc}")
        return False, f"Ошибка обновления раздела: {str(exc)}"

async def ensure_category_exists_in_db(category_name: str) -> Optional[int]:
    """Гарантирует наличие раздела в базе данных."""
    normalized_name = (category_name or '').strip()
    if not normalized_name:
        return None

    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT id FROM categories WHERE lower(name) = lower(?)",
                (normalized_name,)
            ) as cursor:
                row = await cursor.fetchone()
                if row:
                    return row[0]
            try:
                await db.execute(
                    "INSERT INTO categories (name, created_date) VALUES (?, ?)",
                    (normalized_name, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                )
                await db.commit()
                category_id = db.last_insert_rowid()
                logger.info(
                    "📂 Автоматически добавлен раздел из импорта: %s (ID: %s)",
                    normalized_name,
                    category_id,
                )
                return category_id
            except aiosqlite.IntegrityError:
                async with db.execute(
                    "SELECT id FROM categories WHERE lower(name) = lower(?)",
                    (normalized_name,)
                ) as retry_cursor:
                    retry_row = await retry_cursor.fetchone()
                    if retry_row:
                        return retry_row[0]
    except Exception as exc:
        logger.error(f"⚠️ Ошибка гарантии наличия раздела '{normalized_name}': {exc}")
    return None
    
    # ========== ФУНКЦИИ ДЛЯ МАТЕРИАЛОВ ==========
async def init_materials_table():
    """Создает таблицу для материалов склада"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute('''
            CREATE TABLE IF NOT EXISTS materials (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT NOT NULL,
                name TEXT NOT NULL,
                unit TEXT NOT NULL,
                quantity REAL NOT NULL DEFAULT 0,
                unit_cost_without_vat REAL NOT NULL DEFAULT 0,
                total_cost_without_vat REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
        ''')
        await db.commit()

async def init_work_materials_table():
    """Создает таблицу соответствия работ и материалов"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute('''
            CREATE TABLE IF NOT EXISTS work_materials (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                work_id INTEGER NOT NULL,
                material_id INTEGER NOT NULL,
                quantity_per_unit REAL NOT NULL DEFAULT 0,
                UNIQUE(work_id, material_id),
                FOREIGN KEY (work_id) REFERENCES works(id) ON DELETE CASCADE,
                FOREIGN KEY (material_id) REFERENCES materials(id) ON DELETE CASCADE
            )
        ''')
        await db.commit()

async def init_material_history_table():
    """Создает таблицу истории движения материалов"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute('''
            CREATE TABLE IF NOT EXISTS material_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                material_id INTEGER NOT NULL,
                change_type TEXT NOT NULL,
                change_amount REAL NOT NULL,
                resulting_quantity REAL,
                performed_by TEXT,
                description TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (material_id) REFERENCES materials(id) ON DELETE CASCADE
            )
        ''')
        await db.commit()

async def log_material_history_entry(
    db,
    material_id: int,
    change_amount: float,
    change_type: str,
    performed_by: Optional[str] = None,
    description: Optional[str] = None
):
    """Добавляет запись в историю движения материалов"""
    performed_by_value = (performed_by or 'Неизвестно').strip() or 'Неизвестно'
    description_value = (description or '').strip()

    resulting_quantity = None
    async with db.execute(
        "SELECT quantity FROM materials WHERE id = ?",
        (material_id,)
    ) as cursor:
        row = await cursor.fetchone()
        if row is not None:
            resulting_quantity = row[0]

    await db.execute(
        '''INSERT INTO material_history
           (material_id, change_type, change_amount, resulting_quantity, performed_by, description, created_at)
           VALUES (?, ?, ?, ?, ?, ?, ?)''',
        (
            material_id,
            change_type,
            change_amount,
            resulting_quantity,
            performed_by_value,
            description_value,
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
    )

async def get_material_history_from_db(limit: int = 500):
    """Возвращает историю движения материалов"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                '''SELECT mh.id, mh.material_id, m.name, m.unit, mh.change_type, mh.change_amount,
                          mh.resulting_quantity, mh.performed_by, mh.description, mh.created_at
                   FROM material_history mh
                   LEFT JOIN materials m ON m.id = mh.material_id
                   ORDER BY mh.created_at DESC, mh.id DESC
                   LIMIT ?''',
                (limit,)
            ) as cursor:
                rows = await cursor.fetchall()
                history = []
                for row in rows:
                    (entry_id, material_id, material_name, unit, change_type, change_amount,
                     resulting_quantity, performed_by, description, created_at) = row
                    history.append({
                        'id': entry_id,
                        'material_id': material_id,
                        'material_name': material_name,
                        'unit': unit,
                        'change_type': change_type,
                        'change_amount': change_amount,
                        'resulting_quantity': resulting_quantity,
                        'performed_by': performed_by,
                        'description': description,
                        'created_at': created_at
                    })
                return history
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения истории материалов: {e}")
        return []        

async def get_all_materials_from_db():
    """Получает список всех материалов"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT id, category, name, unit, quantity, unit_cost_without_vat, total_cost_without_vat, created_at "
                "FROM materials ORDER BY name"
            ) as cursor:
                rows = await cursor.fetchall()
                materials = []
                for row in rows:
                    (
                        material_id,
                        category,
                        name,
                        unit,
                        quantity,
                        unit_cost_without_vat,
                        total_cost_without_vat,
                        created_at,
                    ) = row
                    materials.append({
                        'id': material_id,
                        'category': category,
                        'name': name,
                        'unit': unit,
                        'quantity': quantity,
                        'unit_cost_without_vat': unit_cost_without_vat or 0,
                        'total_cost_without_vat': total_cost_without_vat or 0,
                        'created_at': created_at
                    })
                logger.info(f"📦 Найдено материалов в БД: {len(materials)}")
                return materials
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения материалов: {e}")
        return []

async def get_material_by_id(material_id: int):
    """Получает материал по ID"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT id, category, name, unit, quantity, unit_cost_without_vat, total_cost_without_vat, created_at "
                "FROM materials WHERE id = ?",
                (material_id,)
            ) as cursor:
                row = await cursor.fetchone()
                if row:
                    (
                        material_id,
                        category,
                        name,
                        unit,
                        quantity,
                        unit_cost_without_vat,
                        total_cost_without_vat,
                        created_at,
                    ) = row
                    return {
                        'id': material_id,
                        'category': category,
                        'name': name,
                        'unit': unit,
                        'quantity': quantity,
                        'unit_cost_without_vat': unit_cost_without_vat or 0,
                        'total_cost_without_vat': total_cost_without_vat or 0,
                        'created_at': created_at
                    }
        return None
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения материала ID {material_id}: {e}")
        return None
    
async def get_material_pricing_from_db(material_id: int) -> Optional[dict]:
    """Возвращает значения стоимости для материала."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT unit_cost_without_vat, total_cost_without_vat FROM materials WHERE id = ?",
                (material_id,),
            ) as cursor:
                row = await cursor.fetchone()
                if row:
                    unit_cost_without_vat, total_cost_without_vat = row
                    return {
                        'unit_cost_without_vat': unit_cost_without_vat or 0,
                        'total_cost_without_vat': total_cost_without_vat or 0,
                    }
    except Exception as exc:
        logger.error(f"⚠️ Ошибка получения стоимости материала ID {material_id}: {exc}")
    return None


async def update_material_pricing_in_db(
    material_id: int,
    unit_cost_without_vat: float,
    total_cost_without_vat: float,
) -> bool:
    """Обновляет значения стоимости для материала."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "UPDATE materials SET unit_cost_without_vat = ?, total_cost_without_vat = ? WHERE id = ?",
                (unit_cost_without_vat, total_cost_without_vat, material_id),
            )
            await db.commit()
            return True
    except Exception as exc:
        logger.error(f"⚠️ Ошибка обновления стоимости материала ID {material_id}: {exc}")
        return False

async def fetch_work_materials_requirements(db, work_id: int):
    """Получает список материалов и норм расхода для указанной работы, используя существующее соединение"""
    async with db.execute('''
        SELECT wm.material_id, wm.quantity_per_unit, m.name, m.unit, m.category, m.quantity
        FROM work_materials wm
        JOIN materials m ON wm.material_id = m.id
        WHERE wm.work_id = ?
    ''', (work_id,)) as cursor:
        rows = await cursor.fetchall()
        materials = []
        for row in rows:
            material_id, quantity_per_unit, name, unit, category, available_quantity = row
            materials.append({
                'material_id': material_id,
                'quantity_per_unit': quantity_per_unit,
                'material_name': name,
                'unit': unit,
                'category': category,
                'available_quantity': available_quantity
            })
        return materials

async def get_work_materials_from_db(work_id: int):
    """Возвращает материалы, закрепленные за работой"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            return await fetch_work_materials_requirements(db, work_id)
    except Exception as e:
        logger.error(f"⚠️ Ошибка получения материалов для работы ID {work_id}: {e}")
        return []

async def get_work_pricing_from_db(work_id: int) -> dict:
    """Возвращает сохраненные значения стоимости для работы."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute(
                "SELECT unit_cost_without_vat, total_cost_without_vat FROM works WHERE id = ?",
                (work_id,),
            ) as cursor:
                row = await cursor.fetchone()
                if row:
                    unit_cost_without_vat, total_cost_without_vat = row
                    return {
                        'unit_cost_without_vat': unit_cost_without_vat or 0,
                        'total_cost_without_vat': total_cost_without_vat or 0,
                    }
    except Exception as exc:
        logger.error(f"⚠️ Ошибка получения стоимости для работы ID {work_id}: {exc}")
    return {'unit_cost_without_vat': 0, 'total_cost_without_vat': 0}


async def update_work_pricing_in_db(work_id: int, unit_cost_without_vat: float, total_cost_without_vat: float) -> bool:
    """Обновляет значения стоимости работы."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute(
                "UPDATE works SET unit_cost_without_vat = ?, total_cost_without_vat = ? WHERE id = ?",
                (unit_cost_without_vat, total_cost_without_vat, work_id),
            )
            await db.commit()
            return True
    except Exception as exc:
        logger.error(f"⚠️ Ошибка обновления стоимости для работы ID {work_id}: {exc}")
        return False

async def replace_work_materials_for_work(work_id: int, materials_data: List[dict]):
    """Полностью заменяет набор материалов для работы"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")
                await db.execute("DELETE FROM work_materials WHERE work_id = ?", (work_id,))

                for item in materials_data:
                    await db.execute(
                        "INSERT INTO work_materials (work_id, material_id, quantity_per_unit) VALUES (?, ?, ?)",
                        (work_id, item['material_id'], item['quantity_per_unit'])
                    )

                await db.commit()
                logger.info(f"🔗 Обновлены материалы для работы ID: {work_id}")
                return True, None
            except aiosqlite.IntegrityError as e:
                await db.rollback()
                logger.error(f"❌ Ошибка целостности при обновлении материалов работы {work_id}: {e}")
                return False, "Невозможно сохранить материалы для работы"
            except Exception as e:
                await db.rollback()
                logger.error(f"⚠️ Ошибка обновления материалов для работы {work_id}: {e}")
                return False, str(e)
    except Exception as e:
        logger.error(f"⚠️ Не удалось установить соединение для обновления материалов работы {work_id}: {e}")
        return False, str(e)


async def insert_material_to_db(material_data: dict, performed_by: Optional[str] = None):
    """Добавляет новый материал"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            unit_cost_without_vat = float(material_data.get('unit_cost_without_vat', 0) or 0)
            total_cost_without_vat = float(material_data.get('total_cost_without_vat', 0) or 0)

            cursor = await db.execute(
                "INSERT INTO materials (category, name, unit, quantity, unit_cost_without_vat, total_cost_without_vat, created_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?)",
                (
                    material_data['category'],
                    material_data['name'],
                    material_data['unit'],
                    material_data['quantity'],
                    unit_cost_without_vat,
                    total_cost_without_vat,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                )
            )
            
            material_id = cursor.lastrowid
            logger.info(f"📦 Добавлен новый материал: {material_data['name']} (ID: {material_id})")
            await log_material_history_entry(
                db,
                material_id,
                material_data['quantity'],
                'Создание',
                performed_by or 'Система',
                'Создание материала'
            )
            await db.commit()
            return material_id
    except Exception as e:
        logger.error(f"⚠️ Ошибка добавления материала {material_data}: {e}")
        raise

async def update_material_in_db(material_id: int, material_data: dict, performed_by: Optional[str] = None):
    """Обновляет материал"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:

            async with db.execute(
                "SELECT quantity FROM materials WHERE id = ?",
                (material_id,)
            ) as cursor:
                row = await cursor.fetchone()
                if not row:
                    return False
                previous_quantity = row[0] if row[0] is not None else 0

            cursor = await db.execute(
                "UPDATE materials SET category = ?, name = ?, unit = ?, quantity = ? WHERE id = ?",
                (
                    material_data['category'],
                    material_data['name'],
                    material_data['unit'],
                    material_data['quantity'],
                    material_id
                )
            )
            row_updated = cursor.rowcount and cursor.rowcount > 0
            if row_updated:
                logger.info(f"📦 Обновлен материал ID: {material_id}")
                change_amount = material_data['quantity'] - previous_quantity
                if abs(change_amount) > 0:
                    await log_material_history_entry(
                        db,
                        material_id,
                        change_amount,
                        'Корректировка',
                        performed_by or 'Система',
                        'Обновление данных материала'
                    )
            await db.commit()
            if row_updated:
                return True
        return False
    except Exception as e:
        logger.error(f"⚠️ Ошибка обновления материала ID {material_id}: {e}")
        return False
    
async def add_quantity_to_material_in_db(
    material_id: int,
    amount: float,
    performed_by: Optional[str] = None,
    description: Optional[str] = None
):
        
    """Увеличивает количество материала на складе"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")

                async with db.execute(
                    "SELECT quantity FROM materials WHERE id = ?",
                    (material_id,)
                ) as cursor:
                    row = await cursor.fetchone()
                    if not row:
                        await db.rollback()
                        return None

                    current_quantity = row[0] if row[0] is not None else 0

                new_quantity = current_quantity + amount

                await db.execute(
                    "UPDATE materials SET quantity = ? WHERE id = ?",
                    (new_quantity, material_id)
                )
                await log_material_history_entry(
                    db,
                    material_id,
                    amount,
                    'Пополнение',
                    performed_by or 'Система',
                    description or 'Пополнение запаса'
                )
                await db.commit()
                logger.info(
                    f"📦 Увеличено количество материала ID: {material_id} на {amount}. Новый остаток: {new_quantity}"
                )
                return new_quantity
            except Exception as inner_error:
                await db.rollback()
                logger.error(
                    f"⚠️ Ошибка при увеличении количества материала ID {material_id}: {inner_error}"
                )
                return None
    except Exception as e:
        logger.error(f"⚠️ Ошибка соединения при увеличении количества материала ID {material_id}: {e}")
        return None

async def delete_material_from_db(material_id: int):
    """Удаляет материал"""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")
                await db.execute("DELETE FROM work_materials WHERE material_id = ?", (material_id,))
                cursor = await db.execute("DELETE FROM materials WHERE id = ?", (material_id,))
                await db.commit()
                if cursor.rowcount and cursor.rowcount > 0:
                    logger.info(f"🗑️ Удален материал ID: {material_id}")
                    return True
            except Exception as inner_error:
                await db.rollback()
                logger.error(f"⚠️ Ошибка при удалении материала ID {material_id}: {inner_error}")
                return False
        return False
    except Exception as e:
        logger.error(f"⚠️ Ошибка удаления материала ID {material_id}: {e}")
        return False

async def delete_category_from_db(category_id: int):
    """Удаляет раздел из базы данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            # Проверяем, используются ли разделы в работах
            async with db.execute(
                "SELECT COUNT(*) FROM works WHERE category = (SELECT name FROM categories WHERE id = ?)", 
                (category_id,)
            ) as cursor:
                usage_count = await cursor.fetchone()
                if usage_count and usage_count[0] > 0:
                    return False, "Нельзя удалить раздел, который используется в работах"
            
            await db.execute("DELETE FROM categories WHERE id = ?", (category_id,))
            await db.commit()
            if db.rowcount > 0:
                logger.info(f"🗑️ Удален раздел ID: {category_id}")
                return True, "Раздел успешно удален"
        return False, "Раздел не найден"
    except Exception as e:
        logger.error(f"⚠️ Ошибка удаления раздела ID {category_id}: {e}")
        return False, f"Ошибка удаления: {str(e)}"

# ========== ФУНКЦИИ ДЛЯ ОТЧЕТОВ ==========
async def get_reports_for_date_from_db(target_date: str):
    """Получает отчеты за конкретную дату из базы данных."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute('''
                SELECT wr.quantity, w.name, w.category, w.unit, f.first_name, f.last_name
                FROM work_reports wr
                JOIN works w ON wr.work_id = w.id
                JOIN foremen f ON wr.foreman_id = f.id
                WHERE wr.report_date = ?
            ''', (target_date,)) as cursor:
                rows = await cursor.fetchall()

            grouped_reports = {}
            for quantity, work_name, category, unit, full_name, position in rows:
                if full_name not in grouped_reports:
                    grouped_reports[full_name] = {
                        'position': position,
                        'works': []
                    }
                grouped_reports[full_name]['works'].append({
                    'name': work_name,
                    'quantity': quantity,
                    'unit': unit
                })

            reports = []
            for foreman, info in grouped_reports.items():
                reports.append({
                    'foreman': foreman,
                    'position': info.get('position'),
                    'works': info['works']                })
            logger.info(f"📊 Найдено отчетов за {target_date}: {len(reports)} бригадиров")
            return reports
    except Exception as e:
        logger.error(f"❌ Ошибка получения отчетов за дату {target_date}: {e}")
        return []

async def get_all_reports_from_db(date_filter=None):
    """Получает все отчеты из базы данных с возможностью фильтрации по дате."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            query = '''
                SELECT wr.id, wr.foreman_id, wr.work_id, wr.quantity,
                       wr.report_date, wr.report_time, wr.photo_report_url,
                       wr.is_verified,
                       f.first_name as foreman_full_name,
                       f.last_name as foreman_position,
                       w.name as work_name, w.unit
                FROM work_reports wr
                LEFT JOIN foremen f ON wr.foreman_id = f.id
                LEFT JOIN works w ON wr.work_id = w.id
            '''
            params = ()
            
            if date_filter:
                query += ' WHERE wr.report_date = ?'
                params = (date_filter,)
                
            query += " ORDER BY datetime(wr.report_date || ' ' || wr.report_time) DESC"
            
            async with db.execute(query, params) as cursor:
                rows = await cursor.fetchall()
                reports = []
                for row in rows:
                    (report_id, foreman_id, work_id, quantity, report_date,
                     report_time, photo_url, is_verified, foreman_full_name,
                     foreman_position, work_name, unit) = row
                    reports.append({
                        'id': report_id,
                        'foreman_id': foreman_id,
                        'work_id': work_id,
                        'quantity': quantity,
                        'report_date': report_date,
                        'report_time': report_time,
                        'photo_report_url': photo_url,
                        'is_verified': bool(is_verified),
                        'foreman_name': foreman_full_name,
                        'foreman_position': foreman_position,
                        'work_name': work_name,
                        'unit': unit
                    })
                logger.info(f"📋 Загружено отчетов: {len(reports)}")
                return reports
    except Exception as e:
        logger.error(f"❌ Ошибка получения всех отчетов: {e}")
        return []

async def get_report_by_id(report_id: int):
    """Получает конкретный отчет по ID."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute('''
                SELECT wr.id, wr.foreman_id, wr.work_id, wr.quantity,
                       wr.report_date, wr.report_time, wr.photo_report_url,
                       wr.is_verified,           
                       f.first_name as foreman_full_name,
                        f.last_name as foreman_position,
                       w.name as work_name, w.unit
                FROM work_reports wr
                LEFT JOIN foremen f ON wr.foreman_id = f.id
                LEFT JOIN works w ON wr.work_id = w.id
                WHERE wr.id = ?
            ''', (report_id,)) as cursor:
                row = await cursor.fetchone()
                if row:
                    (report_id, foreman_id, work_id, quantity, report_date,
                     report_time, photo_url, is_verified, foreman_full_name,
                     foreman_position, work_name, unit) = row
                    return {
                        'id': report_id,
                        'foreman_id': foreman_id,
                        'work_id': work_id,
                        'quantity': quantity,
                        'report_date': report_date,
                        'report_time': report_time,
                        'photo_report_url': photo_url,
                        'is_verified': bool(is_verified),
                        'foreman_name': foreman_full_name,
                        'foreman_position': foreman_position,
                        'work_name': work_name,
                        'unit': unit
                    }
        return None
    except Exception as e:
        logger.error(f"❌ Ошибка получения отчета по ID {report_id}: {e}")
        return None

async def update_report_in_db(report_id: int, report_data: dict):
    """Обновляет отчет в базе данных вместе со всеми связанными остатками."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")

                # Получаем старые данные отчета для восстановления балансов
                async with db.execute(
                    "SELECT work_id, quantity, foreman_id FROM work_reports WHERE id = ?",
                    (report_id,)
                ) as cursor:
                    old_row = await cursor.fetchone()
                    if not old_row:
                        await db.rollback()
                        return False, "Отчет не найден"

                    old_work_id, old_quantity, old_foreman_id = old_row

                # Определяем нового бригадира (если не передан, используем прежнего)
                new_foreman_id_raw = report_data.get('foreman_id', old_foreman_id)
                new_foreman_id = old_foreman_id
                if new_foreman_id_raw is not None:
                    try:
                        new_foreman_id = int(new_foreman_id_raw)
                    except (TypeError, ValueError):
                        await db.rollback()
                        return False, "Некорректный идентификатор бригадира"

                new_foreman_display = await get_foreman_display_name(db, new_foreman_id)
                old_foreman_display = await get_foreman_display_name(db, old_foreman_id)
                correction_display = f"{old_foreman_display} (коррекция отчета ID {report_id})"

                # Восстанавливаем баланс по старой работе
                await db.execute(
                    "UPDATE works SET balance = balance + ? WHERE id = ?",
                    (old_quantity, old_work_id)
                )

                # Возвращаем материалы на склад по старой работе
                old_requirements = await fetch_work_materials_requirements(db, old_work_id)
                for requirement in old_requirements:
                    total_to_restore = requirement['quantity_per_unit'] * old_quantity
                    if total_to_restore <= 0:
                        continue
                    await db.execute(
                        "UPDATE materials SET quantity = quantity + ? WHERE id = ?",
                        (total_to_restore, requirement['material_id'])
                    )
                    await log_material_history_entry(
                        db,
                        requirement['material_id'],
                        total_to_restore,
                        'Возврат',
                        correction_display,
                        f"Возврат при редактировании отчета работы ID {report_id}"
                    )


            # Проверяем наличие работы и доступный баланс под новую работу
                async with db.execute(
                    "SELECT balance FROM works WHERE id = ?",
                    (report_data['work_id'],)
                ) as cursor:
                    new_balance_row = await cursor.fetchone()
                    if not new_balance_row:
                        await db.rollback()
                        return False, "Новая работа не найдена"

                    new_balance = new_balance_row[0]
                    if new_balance < report_data['quantity']:
                        await db.rollback()
                        return False, "Недостаточно материалов на балансе для новой работы"

                new_requirements = await fetch_work_materials_requirements(db, report_data['work_id'])
                for requirement in new_requirements:
                    total_required = requirement['quantity_per_unit'] * report_data['quantity']
                    if total_required <= 0:
                        continue
                    if requirement['available_quantity'] < total_required:
                        await db.rollback()
                        return False, (
                            f"Недостаточно материала \"{requirement['material_name']}\" на складе"
                        )

                auto_photo_url = report_data.get('photo_report_url')
                if not auto_photo_url:
                    auto_photo_url = await ensure_report_folder(
                        db,
                        new_foreman_id,
                        report_data.get('report_date')
                    )

                # Списываем новый объем работ
                await db.execute(
                    "UPDATE works SET balance = balance - ? WHERE id = ?",
                    (report_data['quantity'], report_data['work_id'])
                )

                # Списываем материалы для новой работы
                for requirement in new_requirements:
                    total_required = requirement['quantity_per_unit'] * report_data['quantity']
                    if total_required <= 0:
                        continue
                    await db.execute(
                        "UPDATE materials SET quantity = quantity - ? WHERE id = ?",
                        (total_required, requirement['material_id'])
                    )
                    await log_material_history_entry(
                        db,
                        requirement['material_id'],
                        -total_required,
                        'Списание',
                        new_foreman_display,
                        f"Списание по обновленному отчету работы ID {report_id}"
                    )

                # Обновляем сам отчет
                await db.execute(
                    '''UPDATE work_reports
                       SET foreman_id = ?, work_id = ?, quantity = ?,
                           report_date = ?, report_time = ?, photo_report_url = ?
                       WHERE id = ?''',
                    (new_foreman_id, report_data['work_id'], report_data['quantity'],
                     report_data['report_date'], report_data['report_time'],
                     auto_photo_url or '', report_id)
                )

                await db.commit()
                logger.info(f"📝 Обновлен отчет ID: {report_id}")
                return True, "Отчет успешно обновлен"
            except Exception as e:
                await db.rollback()
                logger.error(f"❌ Ошибка обновления отчета ID {report_id}: {e}")
                return False, f"Ошибка обновления: {str(e)}"
    except Exception as e:
        logger.error(f"❌ Ошибка соединения при обновлении отчета ID {report_id}: {e}")
        return False, "Ошибка соединения с базой данных"

async def delete_report_from_db(report_id: int):
    """Удаляет отчет из базы данных и восстанавливает баланс."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")

                # Получаем данные отчета для восстановления баланса
                async with db.execute(
                    "SELECT work_id, quantity, foreman_id FROM work_reports WHERE id = ?",
                    (report_id,)
                ) as cursor:
                    row = await cursor.fetchone()
                    if not row:
                        return False, "Отчет не найден"

                    work_id, quantity, foreman_id = row

                foreman_display = await get_foreman_display_name(db, foreman_id)
                deletion_display = f"{foreman_display} (удаление отчета ID {report_id})"

                # Восстанавливаем баланс работы
                await db.execute(
                    "UPDATE works SET balance = balance + ? WHERE id = ?",
                    (quantity, work_id)
                )

                # Возвращаем материалы на склад
                requirements = await fetch_work_materials_requirements(db, work_id)
                for requirement in requirements:
                    total_to_restore = requirement['quantity_per_unit'] * quantity
                    if total_to_restore <= 0:
                        continue
                    await db.execute(
                        "UPDATE materials SET quantity = quantity + ? WHERE id = ?",
                        (total_to_restore, requirement['material_id'])
                    )
                    await log_material_history_entry(
                        db,
                        requirement['material_id'],
                        total_to_restore,
                        'Возврат',
                        deletion_display,
                        f"Возврат при удалении отчета работы ID {report_id}"
                    )

                # Удаляем отчет
                await db.execute("DELETE FROM work_reports WHERE id = ?", (report_id,))
                await db.commit()

                logger.info(f"🗑️ Удален отчет ID: {report_id}")
                return True, "Отчет успешно удален"
            except Exception as e:
                await db.rollback()
                raise e
    except Exception as e:
        await db.rollback()
        logger.error(f"❌ Ошибка удаления отчета ID {report_id}: {e}")
        return False, f"Ошибка удаления: {str(e)}"
    
async def set_report_verification_status(report_id: int, is_verified: bool) -> bool:
    """Обновляет флаг проверки отчета."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            cursor = await db.execute(
                "UPDATE work_reports SET is_verified = ? WHERE id = ?",
                (1 if is_verified else 0, report_id)
            )
            await db.commit()
            if cursor.rowcount == 0:
                logger.warning(f"⚠️ Не удалось найти отчет ID {report_id} для обновления статуса проверки")
                return False
            logger.info(
                f"✅ Статус проверки отчета ID {report_id} обновлен на {'проверен' if is_verified else 'не проверен'}"
            )
            return True
    except Exception as exc:
        logger.error(f"❌ Ошибка обновления статуса проверки отчета ID {report_id}: {exc}")
        return False    

async def get_all_work_reports_from_db():
    """Получает все отчеты о работах для вкладки отчетов."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            async with db.execute('''
                SELECT id, foreman_id, work_id, quantity, report_date, report_time, photo_report_url, is_verified
                FROM work_reports
                ORDER BY datetime(report_date || ' ' || report_time) DESC
            ''') as cursor:
                rows = await cursor.fetchall()
                reports = []
                for row in rows:
                    (report_id, foreman_id, work_id, quantity, report_date,
                     report_time, photo_url, is_verified) = row
                    reports.append({
                        'id': report_id,
                        'foreman_id': foreman_id,
                        'work_id': work_id,
                        'quantity': quantity,
                        'report_date': report_date,
                        'report_time': report_time,
                        'photo_report_url': photo_url,
                        'is_verified': bool(is_verified)
                    })
                return reports
    except Exception as e:
        logger.error(f"❌ Ошибка получения всех отчетов: {e}")
        return []

async def create_work_report_in_db(report_data: dict):
    """Создает новый отчет о работе."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")

                foreman_display = await get_foreman_display_name(db, report_data.get('foreman_id'))

                # Проверяем баланс работы
                async with db.execute(
                    "SELECT balance FROM works WHERE id = ?",
                    (report_data['work_id'],)
                ) as cursor:
                    balance_row = await cursor.fetchone()
                    if not balance_row:
                        await db.rollback()
                        return False, "Работа не найдена"

                    balance = balance_row[0]
                    if balance < report_data['quantity']:
                        await db.rollback()
                        return False, "Недостаточно материалов на балансе"

                # Проверяем наличие материалов на складе
                materials_requirements = await fetch_work_materials_requirements(db, report_data['work_id'])
                for requirement in materials_requirements:
                    total_required = requirement['quantity_per_unit'] * report_data['quantity']
                    if total_required <= 0:
                        continue
                    if requirement['available_quantity'] < total_required:
                        await db.rollback()
                        return False, (
                            f"Недостаточно материала \"{requirement['material_name']}\" на складе"
                        )

                # Создаем отчет и получаем его ID

                auto_photo_url = await ensure_report_folder(
                    db,
                    report_data.get('foreman_id'),
                    report_data.get('report_date')
                )
                photo_value = auto_photo_url or report_data.get('photo_report_url', '')

                cursor = await db.execute(
                    '''INSERT INTO work_reports
                       (foreman_id, work_id, quantity, report_date, report_time, photo_report_url)
                       VALUES (?, ?, ?, ?, ?, ?)''',
                    (report_data['foreman_id'], report_data['work_id'], report_data['quantity'],
                     report_data['report_date'], report_data['report_time'],
                     photo_value)
                )
                report_id = cursor.lastrowid


                # Вычитаем из баланса работы
                await db.execute(
                    "UPDATE works SET balance = balance - ? WHERE id = ?",
                    (report_data['quantity'], report_data['work_id'])
                )

                # Вычитаем материалы со склада
                for requirement in materials_requirements:
                    total_required = requirement['quantity_per_unit'] * report_data['quantity']
                    if total_required <= 0:
                        continue
                    await db.execute(
                        "UPDATE materials SET quantity = quantity - ? WHERE id = ?",
                        (total_required, requirement['material_id'])
                    )

                    await log_material_history_entry(
                        db,
                        requirement['material_id'],
                        -total_required,
                        'Списание',
                        foreman_display,
                        f"Списание по отчету работы ID {report_id}"
                    )

                
                await db.commit()
                logger.info(f"📊 Создан отчет ID: {report_id}")
                return True, report_id
            except Exception as e:
                await db.rollback()
                raise e
    except Exception as e:
        logger.error(f"❌ Ошибка создания отчета: {e}")
        return False, f"Ошибка создания: {str(e)}"

async def update_work_report_in_db(report_id: int, report_data: dict):
    """Обновляет отчет о работе."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            try:
                await db.execute("BEGIN")

                # Получаем старые данные
                async with db.execute(
                    "SELECT work_id, quantity, foreman_id FROM work_reports WHERE id = ?",
                    (report_id,)
                ) as cursor:
                    old_row = await cursor.fetchone()
                    if not old_row:
                        await db.rollback()
                        return False, "Отчет не найден"

                    old_work_id, old_quantity, old_foreman_id = old_row

                new_foreman_display = await get_foreman_display_name(db, report_data.get('foreman_id'))
                old_foreman_display = await get_foreman_display_name(db, old_foreman_id)
                correction_display = f"{old_foreman_display} (коррекция отчета ID {report_id})"

                # Восстанавливаем старый баланс работы
                await db.execute(
                    "UPDATE works SET balance = balance + ? WHERE id = ?",
                    (old_quantity, old_work_id)
                )

                # Восстанавливаем материалы на складе
                old_requirements = await fetch_work_materials_requirements(db, old_work_id)
                for requirement in old_requirements:
                    total_to_restore = requirement['quantity_per_unit'] * old_quantity
                    if total_to_restore <= 0:
                        continue
                    await db.execute(
                        "UPDATE materials SET quantity = quantity + ? WHERE id = ?",
                        (total_to_restore, requirement['material_id'])
                    )
                    await log_material_history_entry(
                        db,
                        requirement['material_id'],
                        total_to_restore,
                        'Возврат',
                        correction_display,
                        f"Возврат при редактировании отчета работы ID {report_id}"
                    )

                # Проверяем новый баланс работы
                async with db.execute(
                    "SELECT balance FROM works WHERE id = ?",
                    (report_data['work_id'],)
                ) as cursor:
                    new_balance_row = await cursor.fetchone()
                    if not new_balance_row:
                        await db.rollback()
                        return False, "Новая работа не найдена"

                    new_balance = new_balance_row[0]
                    if new_balance < report_data['quantity']:
                        await db.rollback()
                        return False, "Недостаточно материалов на балансе для новой работы"

                # Проверяем наличие материалов на складе для новой работы
                new_requirements = await fetch_work_materials_requirements(db, report_data['work_id'])
                for requirement in new_requirements:
                    total_required = requirement['quantity_per_unit'] * report_data['quantity']
                    if total_required <= 0:
                        continue
                    if requirement['available_quantity'] < total_required:
                        await db.rollback()
                        return False, (
                            f"Недостаточно материала \"{requirement['material_name']}\" на складе"
                        )

                auto_photo_url = report_data.get('photo_report_url')
                if not auto_photo_url:
                    auto_photo_url = await ensure_report_folder(
                        db,
                        report_data.get('foreman_id'),
                        report_data.get('report_date')
                    )


                # Вычитаем новое количество из баланса работы
                await db.execute(
                    "UPDATE works SET balance = balance - ? WHERE id = ?",
                    (report_data['quantity'], report_data['work_id'])
                )

                # Вычитаем материалы со склада
                for requirement in new_requirements:
                    total_required = requirement['quantity_per_unit'] * report_data['quantity']
                    if total_required <= 0:
                        continue
                    await db.execute(
                        "UPDATE materials SET quantity = quantity - ? WHERE id = ?",
                        (total_required, requirement['material_id'])
                    )
                    await log_material_history_entry(
                        db,
                        requirement['material_id'],
                        -total_required,
                        'Списание',
                        new_foreman_display,
                        f"Списание по обновленному отчету работы ID {report_id}"
                    )

                # Обновляем отчет
                await db.execute(
                    '''UPDATE work_reports
                       SET foreman_id = ?, work_id = ?, quantity = ?,
                           report_date = ?, report_time = ?, photo_report_url = ?
                       WHERE id = ?''',
                    (report_data['foreman_id'], report_data['work_id'], report_data['quantity'],
                     report_data['report_date'], report_data['report_time'],
                     auto_photo_url or '', report_id)
                )

                await db.commit()
                logger.info(f"📊 Обновлен отчет ID: {report_id}")
                return True, "Отчет успешно обновлен"
            except Exception as e:
                await db.rollback()
                raise e
    except Exception as e:
        logger.error(f"❌ Ошибка обновления отчета ID {report_id}: {e}")
        return False, f"Ошибка обновления: {str(e)}"

# ========== ЭНДПОИНТЫ API ==========
@app.get("/")
def read_root():
    return {"message": "StroyKontrol API", "version": "1.0.0"}

# Таблица пользователей сайта
async def init_site_users_table():
    """Создает таблицу для пользователей сайта"""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute('''
            CREATE TABLE IF NOT EXISTS site_users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                is_active INTEGER NOT NULL DEFAULT 1,
                created_date TEXT NOT NULL,
                last_login TEXT
            )
        ''')
        await db.commit()

# Инициализируем таблицу при запуске
@app.on_event("startup")
async def startup_event():
    await init_site_users_table()
    await init_categories_table()
    await init_materials_table()
    await init_work_materials_table()
    await init_material_history_table()
    await ensure_work_reports_verification_column()
    await ensure_work_pricing_columns()
    await ensure_material_pricing_columns()


@app.get("/api/works/export")
async def export_works():
    """Экспорт работ в Excel файл"""
    try:
        works = await get_all_works_from_db()
        
        # Создаем новую книгу Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Работы"
        
        # Заголовки
        headers = ["ID", "Название работы", "Раздел", "Единица измерения", "На балансе", "Проект", "Активна"]
        ws.append(headers)
        
        # Данные
        for work in works:
            row = [
                work['id'],
                work['Название работы'],
                work['Раздел'],
                work['Единица измерения'],
                work['На балансе'],
                work['Проект'],
                "Да" if work['is_active'] else "Нет"
            ]
            ws.append(row)
        
        # Сохраняем в байтовый поток
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        # Возвращаем файл
        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=works_export.xlsx"}
        )
    except Exception as e:
        logger.error(f"❌ Ошибка экспорта работ: {e}")
        raise HTTPException(status_code=500, detail="Ошибка экспорта работ")

# Импорт/экспорт       

@app.post("/api/works/import")
async def import_works(file: UploadFile = File(...)):
    """Импорт работ из Excel файла"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Поддерживаются только файлы Excel (.xlsx, .xls)")
        
        # Читаем файл
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        imported_count = 0
        updated_count = 0
        errors = []
        
        # Пропускаем заголовок (первая строка)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):  # Пропускаем пустые строки
                continue
                
            try:
                work_id, name, category, unit, balance, project_total, is_active_str = row
                
                # Валидация обязательных полей
                if not name or not category or not unit:
                    errors.append(f"Пропущены обязательные поля в строке: {row}")
                    continue
                
                # Преобразуем данные
                normalized_category = str(category).strip()
                await ensure_category_exists_in_db(normalized_category)

                work_data = {
                    'name': str(name).strip(),
                    'category': normalized_category,
                    'unit': str(unit).strip(),
                    'balance': float(balance) if balance else 0,
                    'project_total': float(project_total) if project_total else 0,
                    'is_active': 1 if str(is_active_str).lower() in ['да', 'yes', 'true', '1'] else 0
                }
                
                if work_id and await get_work_by_id(int(work_id)):
                    # Обновляем существующую работу
                    success = await update_work_in_db(int(work_id), work_data)
                    if success:
                        updated_count += 1
                    else:
                        errors.append(f"Ошибка обновления работы ID {work_id}")
                else:
                    # Создаем новую работу
                    new_id = await insert_work_to_db(work_data)
                    if new_id:
                        imported_count += 1
                    else:
                        errors.append(f"Ошибка создания работы: {work_data['name']}")
                        
            except Exception as e:
                errors.append(f"Ошибка обработки строки {row}: {str(e)}")
        
        return {
            "success": True,
            "message": f"Импорт завершен. Добавлено: {imported_count}, обновлено: {updated_count}",
            "errors": errors,
            "imported_count": imported_count,
            "updated_count": updated_count,
            "error_count": len(errors)
        }
        
    except Exception as e:
        logger.error(f"❌ Ошибка импорта работ: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка импорта работ: {str(e)}")


@app.get("/api/materials/export")
async def export_materials():
    """Экспорт материалов в Excel файл"""
    try:
        materials = await get_all_materials_from_db()

        wb = Workbook()
        ws = wb.active
        ws.title = "Материалы"

        headers = ["ID", "Раздел", "Название материала", "Единица измерения", "Количество"]
        ws.append(headers)

        for material in materials:
            row = [
                material['id'],
                material['category'],
                material['name'],
                material['unit'],
                material['quantity']
            ]
            ws.append(row)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=materials_export.xlsx"}
        )
    except Exception as e:
        logger.error(f"❌ Ошибка экспорта материалов: {e}")
        raise HTTPException(status_code=500, detail="Ошибка экспорта материалов")


@app.get("/api/materials/template")
async def download_materials_template():
    """Скачивание шаблона Excel для материалов"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Материалы"

        headers = ["ID", "Раздел", "Название материала", "Единица измерения", "Количество"]
        ws.append(headers)
        ws.append(["", "Раздел", "Пример материала", "шт", 0])

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return StreamingResponse(
            file_stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=materials_template.xlsx"}
        )
    except Exception as e:
        logger.error(f"❌ Ошибка формирования шаблона материалов: {e}")
        raise HTTPException(status_code=500, detail="Ошибка формирования шаблона")


@app.post("/api/materials/import")
async def import_materials(file: UploadFile = File(...)):
    """Импорт материалов из Excel"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Поддерживаются только файлы Excel (.xlsx, .xls)")

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active

        imported_count = 0
        updated_count = 0
        errors = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            try:
                material_id, category, name, unit, quantity = row

                if not category or not name or not unit:
                    errors.append(f"Пропущены обязательные поля в строке: {row}")
                    continue

                material_payload = {
                    'category': str(category).strip(),
                    'name': str(name).strip(),
                    'unit': str(unit).strip(),
                    'quantity': float(quantity) if quantity is not None else 0
                }

                if material_id:
                    existing_material = await get_material_by_id(int(material_id))
                    if existing_material:
                        success = await update_material_in_db(int(material_id), material_payload)
                        if success:
                            updated_count += 1
                        else:
                            errors.append(f"Ошибка обновления материала ID {material_id}")
                        continue

                await insert_material_to_db(material_payload)
                imported_count += 1

            except Exception as exc:
                errors.append(f"Ошибка обработки строки {row}: {str(exc)}")

        return {
            "success": True,
            "message": f"Импорт завершен. Добавлено: {imported_count}, обновлено: {updated_count}",
            "errors": errors,
            "imported_count": imported_count,
            "updated_count": updated_count,
            "error_count": len(errors)
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка импорта материалов: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка импорта материалов: {str(e)}")

# Эндпоинты аутентификации

@app.post("/api/site-login")
@app.post("//api/site-login")
async def login_site_user(request: Request):
    """Аутентификация пользователя сайта"""
    try:
        login_data = await request.json()
        logger.info(f"🔐 Попытка входа: username={login_data.get('username')}")
        
        required_fields = ["username", "password"]
        for field in required_fields:
            if field not in login_data:
                logger.error(f"❌ Отсутствует поле: {field}")
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")
        
        password_hash = hash_password(login_data['password'])
        logger.info(f"🔐 Хэш пароля: {password_hash}")
        
        async with aiosqlite.connect(DB_PATH) as db:
            logger.info(f"🔍 Поиск пользователя в БД: {login_data['username']}")
            
            async with db.execute(
                "SELECT id, username, role, is_active FROM site_users WHERE username = ? AND password_hash = ? AND is_active = 1",
                (login_data['username'], password_hash)
            ) as cursor:
                user = await cursor.fetchone()
                logger.info(f"🔍 Результат запроса: {user}")
                
                if user:
                    user_id, username, role, is_active = user
                    logger.info(f"✅ Успешный вход: {username} (id={user_id})")
                    
                    # Обновляем время последнего входа
                    await db.execute(
                        "UPDATE site_users SET last_login = ? WHERE id = ?",
                        (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user_id)
                    )
                    await db.commit()
                    
                    return {
                        "success": True, 
                        "message": "Успешный вход",
                        "user": {
                            "id": user_id,
                            "username": username,
                            "role": role
                        }
                    }
                else:
                    logger.error(f"❌ Неверные учетные данные: username={login_data['username']}")
                    raise HTTPException(status_code=401, detail="Неверное имя пользователя или пароль")
                    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка входа: {e}")
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=500, detail="Ошибка входа")
                    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка входа: {e}")
        raise HTTPException(status_code=500, detail="Ошибка входа")

# ========== ЭНДПОИНТЫ ДЛЯ КАТЕГОРИЙ ==========
@app.get("/api/categories")
async def get_categories():
    """Получает все разделы."""
    categories = await get_categories_from_db()
    return {"success": True, "data": categories}

@app.post("/api/categories")
async def create_category(request: Request):
    """Создает новый раздел."""
    try:
        category_data = await request.json()
        
        if 'name' not in category_data or not category_data['name'].strip():
            raise HTTPException(status_code=400, detail="Отсутствует название раздела")
        
        category_id = await create_category_in_db(category_data)
        if category_id is not None:
            return {"success": True, "message": "Раздел успешно добавлен", "data": {"id": category_id}}
        else:
            raise HTTPException(status_code=500, detail="Ошибка при добавлении раздела в БД")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка при создании раздела: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.put("/api/categories/{category_id}")
async def update_category(category_id: int, request: Request):
    """Обновляет существующий раздел."""
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")

    if 'name' not in payload or not isinstance(payload['name'], str):
        raise HTTPException(status_code=400, detail="Отсутствует название раздела")

    normalized_name = payload['name'].strip()
    if not normalized_name:
        raise HTTPException(status_code=400, detail="Отсутствует название раздела")
    success, message = await update_category_in_db(category_id, normalized_name)

    if success:
        return {
            "success": True,
            "message": message,
            "data": {"id": category_id, "name": normalized_name}
        }

    if message in {"Отсутствует название раздела", "Раздел с таким названием уже существует"}:
        raise HTTPException(status_code=400, detail=message)
    if message == "Раздел не найден":
        raise HTTPException(status_code=404, detail=message)

    raise HTTPException(status_code=500, detail=message or "Не удалось обновить раздел")

@app.delete("/api/categories/{category_id}")
async def delete_category(category_id: int):
    """Удаляет раздел."""
    success, message = await delete_category_from_db(category_id)
    if success:
        return {"success": True, "message": message}
    else:
        raise HTTPException(status_code=400, detail=message)

#Обновляем функцию startup_event для инициализации таблицы разделов
@app.on_event("startup")
async def startup_event():
    await init_site_users_table()
    await init_categories_table()
    await init_materials_table()
    await init_work_materials_table()
    await ensure_work_reports_verification_column()
    await ensure_work_pricing_columns()
    await ensure_material_pricing_columns()


# Эндпоинты для работ

@app.get("/api/all-works")
async def get_all_works():
    """Получает ВСЕ работы (для сайта)."""
    works = await get_all_works_from_db()
    return {"success": True, "data": works}

@app.get("/api/works")
async def get_works():
    works = await get_active_works_from_db()
    return {"success": True, "data": works}

@app.get("/api/works/{work_id}")
async def get_work(work_id: int):
    work = await get_work_by_id(work_id)
    if work is None:
        raise HTTPException(status_code=404, detail="Работа не найдена")
    return {"success": True, "data": work}

@app.post("/api/works")
async def create_work(request: Request):
    try:
        work_data = await request.json()
        logger.info(f"DEBUG: create_work получил данные: {work_data}")

        # Валидация данных
        required_fields = ["name", "category", "unit", "balance", "is_active"]
        for field in required_fields:
            if field not in work_data:
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")
        if not isinstance(work_data['is_active'], int) or work_data['is_active'] not in [0, 1]:
             raise HTTPException(status_code=400, detail="is_active должно быть 0 или 1")
        if not isinstance(work_data['balance'], int) or work_data['balance'] < 0:
             raise HTTPException(status_code=400, detail="balance должно быть числом >= 0")

        # Добавляем project_total по умолчанию, если не указан
        if 'project_total' not in work_data:
            work_data['project_total'] = 0

        try:
            unit_cost_without_vat = float(work_data.get('unit_cost_without_vat', 0) or 0)
            total_cost_without_vat = float(work_data.get('total_cost_without_vat', 0) or 0)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Стоимость должна быть числом")

        if unit_cost_without_vat < 0 or total_cost_without_vat < 0:
            raise HTTPException(status_code=400, detail="Стоимость не может быть отрицательной")

        work_data['unit_cost_without_vat'] = unit_cost_without_vat
        work_data['total_cost_without_vat'] = total_cost_without_vat
   

        work_id = await insert_work_to_db(work_data)
        if work_id is not None:
            created_work = await get_work_by_id(work_id)
            return {"success": True, "message": "Работа успешно добавлена", "data": created_work}
        else:
            raise HTTPException(status_code=500, detail="Ошибка при добавлении работы в БД")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка при создании работы: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.put("/api/works/{work_id}")
async def update_work(work_id: int, request: Request):
     # Проверяем, существует ли работа
    existing_work = await get_work_by_id(work_id)
    if existing_work is None:
        raise HTTPException(status_code=404, detail="Работа не найдена для обновления")
    try:
        work_data = await request.json()
        # Валидация данных
        required_fields = ["name", "category", "unit", "balance", "is_active"]
        for field in required_fields:
            if field not in work_data:
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")
        if not isinstance(work_data['is_active'], int) or work_data['is_active'] not in [0, 1]:
             raise HTTPException(status_code=400, detail="is_active должно быть 0 или 1")
        if not isinstance(work_data['balance'], (int, float)) or work_data['balance'] < 0:
             raise HTTPException(status_code=400, detail="balance должно быть числом >= 0")

        # Добавляем project_total по умолчанию, если не указан
        if 'project_total' not in work_data:
            work_data['project_total'] = 0

        try:
            unit_cost_without_vat = float(work_data.get('unit_cost_without_vat', 0) or 0)
            total_cost_without_vat = float(work_data.get('total_cost_without_vat', 0) or 0)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Стоимость должна быть числом")

        if unit_cost_without_vat < 0 or total_cost_without_vat < 0:
            raise HTTPException(status_code=400, detail="Стоимость не может быть отрицательной")

        work_data['unit_cost_without_vat'] = unit_cost_without_vat
        work_data['total_cost_without_vat'] = total_cost_without_vat

        success = await update_work_in_db(work_id, work_data)
        if success:
            updated_work = await get_work_by_id(work_id)
            return {"success": True, "message": "Работа успешно обновлена", "data": updated_work}
        else:
            raise HTTPException(status_code=500, detail="Ошибка при обновлении работы в БД")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка при обновлении работы ID {work_id}: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.put("/api/works/{work_id}/add-balance")
async def add_work_balance(work_id: int, request: Request):
    existing_work = await get_work_by_id(work_id)
    if existing_work is None:
        raise HTTPException(status_code=404, detail="Работа не найдена")

    try:
        payload = await request.json()
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")

    pricing_payload = None

    if 'amount' not in payload:
        raise HTTPException(status_code=400, detail="Отсутствует поле: amount")

    try:
        amount = float(payload['amount'])
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="amount должно быть числом")

    if amount <= 0:
        raise HTTPException(status_code=400, detail="amount должно быть больше 0")

    new_balance = await add_balance_to_work_in_db(work_id, amount)
    if new_balance is None:
        raise HTTPException(status_code=500, detail="Не удалось обновить баланс работы")

    existing_work['На балансе'] = new_balance
    existing_work['balance'] = new_balance

    return {
        "success": True,
        "message": "Баланс работы успешно обновлен",
        "data": existing_work
    }

@app.delete("/api/works/{work_id}")
async def delete_work(work_id: int):
    # Проверяем, существует ли работа
    existing_work = await get_work_by_id(work_id)
    if existing_work is None:
        raise HTTPException(status_code=404, detail="Работа не найдена для удаления")

    success = await delete_work_from_db(work_id)
    if success:
        return {"success": True, "message": "Работа успешно удалена"}
    else:
        raise HTTPException(status_code=500, detail="Ошибка при удалении работы из БД")


@app.get("/api/works/{work_id}/materials")
async def get_work_materials(work_id: int):
    work = await get_work_by_id(work_id)
    if work is None:
        raise HTTPException(status_code=404, detail="Работа не найдена")

    materials_for_work = await get_work_materials_from_db(work_id)
    pricing = await get_work_pricing_from_db(work_id)
    return {"success": True, "data": materials_for_work, "pricing": pricing}

@app.put("/api/works/{work_id}/materials")
async def update_work_materials(work_id: int, request: Request):
    work = await get_work_by_id(work_id)
    if work is None:
        raise HTTPException(status_code=404, detail="Работа не найдена")

    try:
        payload = await request.json()
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")

    pricing_payload = None

    if isinstance(payload, dict):
        pricing_payload = payload.get('pricing')

        if 'materials' in payload:
            materials_list = payload['materials']
        else:
            materials_list = []
    elif isinstance(payload, list):
        materials_list = payload
    elif payload is None:
        materials_list = []
    else:
        raise HTTPException(status_code=400, detail="Некорректный формат данных")

    normalized_materials = []
    seen_ids = set()

    for item in materials_list:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="Элемент материалов должен быть объектом")
        if 'material_id' not in item or 'quantity_per_unit' not in item:
            raise HTTPException(status_code=400, detail="Отсутствуют обязательные поля")

        try:
            material_id = int(item['material_id'])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="material_id должен быть целым числом")

        try:
            quantity_per_unit = float(item['quantity_per_unit'])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="quantity_per_unit должен быть числом")

        if material_id <= 0:
            raise HTTPException(status_code=400, detail="material_id должен быть положительным")
        if quantity_per_unit < 0:
            raise HTTPException(status_code=400, detail="Количество не может быть отрицательным")
        if material_id in seen_ids:
            raise HTTPException(status_code=400, detail="Материал не может повторяться")

        material = await get_material_by_id(material_id)
        if material is None:
            raise HTTPException(status_code=404, detail=f"Материал ID {material_id} не найден")

        seen_ids.add(material_id)

        if quantity_per_unit == 0:
            continue

        normalized_materials.append({
            'material_id': material_id,
            'quantity_per_unit': quantity_per_unit
        })

    unit_cost_without_vat = None
    total_cost_without_vat = None

    if pricing_payload is not None:
        if not isinstance(pricing_payload, dict):
            raise HTTPException(status_code=400, detail="Некорректный формат данных стоимости")
        try:
            unit_cost_without_vat = float(pricing_payload.get('unit_cost_without_vat', 0) or 0)
            total_cost_without_vat = float(pricing_payload.get('total_cost_without_vat', 0) or 0)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Стоимость должна быть числом")

        if unit_cost_without_vat < 0 or total_cost_without_vat < 0:
            raise HTTPException(status_code=400, detail="Стоимость не может быть отрицательной")

    success, error_message = await replace_work_materials_for_work(work_id, normalized_materials)
    if not success:
        raise HTTPException(status_code=400, detail=error_message or "Не удалось сохранить материалы")
    
    if unit_cost_without_vat is not None and total_cost_without_vat is not None:
        pricing_saved = await update_work_pricing_in_db(work_id, unit_cost_without_vat, total_cost_without_vat)
        if not pricing_saved:
            raise HTTPException(status_code=500, detail="Не удалось сохранить стоимость работы")

    updated_materials = await get_work_materials_from_db(work_id)
    updated_pricing = await get_work_pricing_from_db(work_id)
    return {
        "success": True,
        "message": "Материалы для работы обновлены",
        "data": updated_materials,
        "pricing": updated_pricing
    }

# Эндпоинты для материалов склада
@app.get("/api/materials")
async def get_materials():
    materials = await get_all_materials_from_db()
    return {"success": True, "data": materials}

@app.get("/api/materials/history")
async def get_material_history(limit: int = 500):
    history = await get_material_history_from_db(limit)
    return {"success": True, "data": history}

@app.get("/api/materials/{material_id}")
async def get_material(material_id: int):
    material = await get_material_by_id(material_id)
    if material is None:
        raise HTTPException(status_code=404, detail="Материал не найден")
    return {"success": True, "data": material}

@app.get("/api/materials/{material_id}/pricing")
async def get_material_pricing(material_id: int):
    material = await get_material_by_id(material_id)
    if material is None:
        raise HTTPException(status_code=404, detail="Материал не найден")

    pricing = await get_material_pricing_from_db(material_id)
    if pricing is None:
        pricing = {'unit_cost_without_vat': 0, 'total_cost_without_vat': 0}

    return {"success": True, "data": pricing}


@app.put("/api/materials/{material_id}/pricing")
async def update_material_pricing(material_id: int, request: Request):
    material = await get_material_by_id(material_id)
    if material is None:
        raise HTTPException(status_code=404, detail="Материал не найден")

    try:
        payload = await request.json()
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Неверный формат данных")

    def parse_cost(value, field_name):
        if value is None or value == "":
            return 0.0
        try:
            numeric = float(value)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail=f"{field_name} должно быть числом")
        if numeric < 0:
            raise HTTPException(status_code=400, detail=f"{field_name} должно быть >= 0")
        return float(numeric)

    unit_cost_without_vat = parse_cost(payload.get('unit_cost_without_vat'), 'unit_cost_without_vat')
    total_cost_without_vat = parse_cost(payload.get('total_cost_without_vat'), 'total_cost_without_vat')

    updated = await update_material_pricing_in_db(
        material_id,
        unit_cost_without_vat,
        total_cost_without_vat,
    )

    if not updated:
        raise HTTPException(status_code=500, detail="Не удалось сохранить стоимость материала")

    updated_pricing = await get_material_pricing_from_db(material_id) or {
        'unit_cost_without_vat': unit_cost_without_vat,
        'total_cost_without_vat': total_cost_without_vat,
    }

    return {
        "success": True,
        "message": "Стоимость материала обновлена",
        "data": updated_pricing,
    }


@app.post("/api/materials")
async def create_material(request: Request):
    try:
        material_data = await request.json()
        performed_by = material_data.pop('performed_by', None)
        

        required_fields = ["name", "category", "unit", "quantity"]
        for field in required_fields:
            if field not in material_data or (isinstance(material_data[field], str) and not material_data[field].strip()):
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")

        try:
            material_data['quantity'] = float(material_data['quantity'])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="quantity должно быть числом")

        if material_data['quantity'] < 0:
            raise HTTPException(status_code=400, detail="quantity должно быть >= 0")

        material_id = await insert_material_to_db(material_data, performed_by or 'Система')
        created_material = await get_material_by_id(material_id)
        return {"success": True, "message": "Материал успешно добавлен", "data": created_material}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка при создании материала: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.put("/api/materials/{material_id}/add-quantity")
async def add_material_quantity_endpoint(material_id: int, request: Request):
    material = await get_material_by_id(material_id)
    if material is None:
        raise HTTPException(status_code=404, detail="Материал не найден")

    try:
        payload = await request.json()
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")

    if 'amount' not in payload:
        raise HTTPException(status_code=400, detail="Отсутствует поле: amount")

    try:
        amount = float(payload['amount'])
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="amount должно быть числом")

    if amount <= 0:
        raise HTTPException(status_code=400, detail="amount должно быть больше 0")

    performed_by = payload.get('performed_by')
    description = payload.get('description')

    new_quantity = await add_quantity_to_material_in_db(
        material_id,
        amount,
        performed_by or 'Система',
        description
    )
    if new_quantity is None:
        raise HTTPException(status_code=500, detail="Не удалось обновить количество материала")

    material['quantity'] = new_quantity

    return {
        "success": True,
        "message": "Количество материала успешно обновлено",
        "data": material
    }

@app.put("/api/materials/{material_id}")
async def update_material(material_id: int, request: Request):
    existing_material = await get_material_by_id(material_id)
    if existing_material is None:
        raise HTTPException(status_code=404, detail="Материал не найден для обновления")

    try:
        material_data = await request.json()

        required_fields = ["name", "category", "unit", "quantity"]
        for field in required_fields:
            if field not in material_data or (isinstance(material_data[field], str) and not material_data[field].strip()):
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")

        try:
            material_data['quantity'] = float(material_data['quantity'])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="quantity должно быть числом")

        if material_data['quantity'] < 0:
            raise HTTPException(status_code=400, detail="quantity должно быть >= 0")

        success = await update_material_in_db(material_id, material_data, performed_by or 'Система')
        if success:
            updated_material = await get_material_by_id(material_id)
            return {"success": True, "message": "Материал успешно обновлен", "data": updated_material}
        else:
            raise HTTPException(status_code=500, detail="Ошибка при обновлении материала в БД")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка при обновлении материала ID {material_id}: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")


@app.delete("/api/materials/{material_id}")
async def delete_material(material_id: int):
    existing_material = await get_material_by_id(material_id)
    if existing_material is None:
        raise HTTPException(status_code=404, detail="Материал не найден для удаления")

    success = await delete_material_from_db(material_id)
    if success:
        return {"success": True, "message": "Материал успешно удален"}
    else:
        raise HTTPException(status_code=500, detail="Ошибка при удалении материала из БД")

# Эндпоинты для бригадиров
@app.get("/api/foremen")
async def get_foremen():
    foremen = await get_foremen_from_db()
    return {"success": True, "data": foremen}

@app.post("/api/foremen")
async def create_foreman(request: Request):
    """Создает нового бригадира."""
    try:
        foreman_data = await request.json()

        # Поддержка старых названий полей
        if 'full_name' not in foreman_data and 'first_name' in foreman_data:
            foreman_data['full_name'] = foreman_data['first_name']
        if 'position' not in foreman_data and 'last_name' in foreman_data:
            foreman_data['position'] = foreman_data['last_name']
        
        # Валидация
        required_fields = ["full_name", "position"]
        for field in required_fields:
            if field not in foreman_data:
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")
        
        foreman_id = await create_foreman_in_db(foreman_data)
        if foreman_id is not None:
            return {"success": True, "message": "Бригадир успешно добавлен", "data": {"id": foreman_id}}
        else:
            raise HTTPException(status_code=500, detail="Ошибка при добавлении бригадира в БД")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")
    except Exception as e:
        logger.error(f"❌ Ошибка при создании бригадира: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

# В эндпоинте update_foreman добавим поддержку is_active
@app.put("/api/foremen/{foreman_id}")
async def update_foreman(foreman_id: int, request: Request):
    """Обновляет данные бригадира."""
    try:
        foreman_data = await request.json()

        # Поддержка старых названий полей
        if 'full_name' not in foreman_data and 'first_name' in foreman_data:
            foreman_data['full_name'] = foreman_data['first_name']
        if 'position' not in foreman_data and 'last_name' in foreman_data:
            foreman_data['position'] = foreman_data['last_name']
        
        # Валидация
        required_fields = ["full_name", "position"]
        for field in required_fields:
            if field not in foreman_data:
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")
        
        if 'is_active' in foreman_data:
            is_active_value = foreman_data['is_active']
            if isinstance(is_active_value, bool):
                is_active_value = int(is_active_value)
            if not isinstance(is_active_value, int):
                raise HTTPException(status_code=400, detail="is_active должно быть числом 0 или 1")
            if is_active_value not in [0, 1]:
                raise HTTPException(status_code=400, detail="is_active должно быть 0 или 1")
            foreman_data['is_active'] = is_active_value
        
        success = await update_foreman_in_db(foreman_id, foreman_data)
        if success:
            return {"success": True, "message": "Бригадир успешно обновлен"}
        else:
            raise HTTPException(status_code=500, detail="Ошибка при обновлении бригадира в БД")
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")
    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        logger.error(f"❌ Ошибка при обновлении бригадира: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.delete("/api/foremen/{foreman_id}")
async def delete_foreman(foreman_id: int):
    """Удаляет бригадира."""
    success, message = await delete_foreman_from_db(foreman_id)
    if success:
        return {"success": True, "message": message}
    else:
        raise HTTPException(status_code=400, detail=message)
    
@app.get("/api/foremen/{foreman_id}/sections")
async def get_foreman_sections(foreman_id: int):
    if not await foreman_exists(foreman_id):
        raise HTTPException(status_code=404, detail="Бригадир не найден")

    sections = await get_foreman_sections_from_db(foreman_id)
    return {"success": True, "data": sections}


@app.put("/api/foremen/{foreman_id}/sections")
async def update_foreman_sections(foreman_id: int, request: Request):
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")

    if isinstance(payload, dict):
        raw_ids = payload.get('category_ids')
        if raw_ids is None:
            raw_ids = payload.get('categories', [])
    elif isinstance(payload, list):
        raw_ids = payload
    elif payload is None:
        raw_ids = []
    else:
        raise HTTPException(status_code=400, detail="Некорректный формат данных")

    category_ids = []
    seen = set()
    for item in raw_ids:
        try:
            category_id = int(item)
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="Идентификатор раздела должен быть числом")
        if category_id <= 0:
            raise HTTPException(status_code=400, detail="Идентификатор раздела должен быть положительным")
        if category_id not in seen:
            seen.add(category_id)
            category_ids.append(category_id)

    success, message, sections = await replace_foreman_sections_for_foreman(foreman_id, category_ids)
    if not success:
        if message == "Бригадир не найден":
            raise HTTPException(status_code=404, detail=message)
        raise HTTPException(status_code=400, detail=message or "Не удалось сохранить разделы")

    return {
        "success": True,
        "message": "Разделы бригадира обновлены",
        "data": sections or []
    }

# Эндпоинты для отчетов
@app.get("/api/reports/{date}")
async def get_reports(date: str):
    # Проверка формата даты
    try:
        datetime.strptime(date, '%Y-%m-%d')
    except ValueError:
        raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD")

    reports = await get_reports_for_date_from_db(date)
    return {"success": True, "data": reports}

@app.get("/api/all-reports")
async def get_all_reports(date: Optional[str] = None):
    """Получает все отчеты с возможностью фильтрации по дате."""
    reports = await get_all_reports_from_db(date)
    return {"success": True, "data": reports}

@app.get("/api/report/{report_id}")
async def get_report(report_id: int):
    """Получает конкретный отчет по ID."""
    report = await get_report_by_id(report_id)
    if report is None:
        raise HTTPException(status_code=404, detail="Отчет не найден")
    return {"success": True, "data": report}

@app.put("/api/report/{report_id}")
async def update_report(report_id: int, request: Request):
    """Обновляет существующий отчет."""
    # Проверяем, существует ли отчет
    existing_report = await get_report_by_id(report_id)
    if existing_report is None:
        raise HTTPException(status_code=404, detail="Отчет не найден")
    
    try:
        report_data = await request.json()
        
        if isinstance(report_data, dict) and isinstance(report_data.get('works'), list):
            base_required = ["foreman_id", "report_date", "report_time"]
            for field in base_required:
                if field not in report_data:
                    raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")

            works_list = [item for item in report_data.get('works', []) if item is not None]
            if not works_list:
                raise HTTPException(status_code=400, detail="Добавьте хотя бы одну работу в отчет")

            created_ids = []
            for work_item in works_list:
                if not isinstance(work_item, dict):
                    raise HTTPException(status_code=400, detail="Неверный формат данных работы")

                for key in ("work_id", "quantity"):
                    if key not in work_item:
                        raise HTTPException(status_code=400, detail="Каждая работа должна содержать идентификатор и количество")

                try:
                    work_id_value = int(work_item["work_id"])
                except (TypeError, ValueError):
                    raise HTTPException(status_code=400, detail="Некорректный идентификатор работы")

                try:
                    quantity_value = float(work_item["quantity"])
                except (TypeError, ValueError):
                    raise HTTPException(status_code=400, detail="Некорректное количество выполненной работы")

                if quantity_value <= 0:
                    raise HTTPException(status_code=400, detail="Количество выполненной работы должно быть больше нуля")

                payload = {
                    "foreman_id": report_data["foreman_id"],
                    "work_id": work_id_value,
                    "quantity": quantity_value,
                    "report_date": report_data["report_date"],
                    "report_time": report_data["report_time"],
                }

                if "photo_report_url" in report_data:
                    payload["photo_report_url"] = report_data["photo_report_url"]

                success, result = await create_work_report_in_db(payload)
                if not success:
                    raise HTTPException(status_code=400, detail=result)

                created_ids.append(result)

            return {
                "success": True,
                "message": f"Создано {len(created_ids)} отчетов" if len(created_ids) > 1 else "Отчет успешно создан",
                "data": {"ids": created_ids}
            }

        # Валидация для одиночного формата
        required_fields = ["work_id", "quantity", "report_date", "report_time"]
        for field in required_fields:
            if field not in report_data:
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")
        
        if not isinstance(report_data['work_id'], int) or report_data['work_id'] <= 0:
            raise HTTPException(status_code=400, detail="work_id должен быть положительным целым числом")
        
        if not isinstance(report_data['quantity'], (int, float)) or report_data['quantity'] <= 0:
            raise HTTPException(status_code=400, detail="quantity должен быть положительным числом")
        
        success, message = await update_report_in_db(report_id, report_data)
        if success:
            updated_report = await get_report_by_id(report_id)
            return {"success": True, "message": message, "data": updated_report}
        else:
            raise HTTPException(status_code=400, detail=message)
            
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Ошибка при обновлении отчета ID {report_id}: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.post("/api/report/{report_id}/verify")
async def verify_report(report_id: int, request: Request):
    """Обновляет статус проверки отчета."""
    try:
        payload = await request.json()
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")

    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Ожидается JSON-объект")

    if 'is_verified' not in payload:
        raise HTTPException(status_code=400, detail="Отсутствует поле is_verified")

    raw_value = payload['is_verified']

    if isinstance(raw_value, bool):
        desired_status = raw_value
    elif isinstance(raw_value, int):
        desired_status = bool(raw_value)
    elif isinstance(raw_value, str):
        normalized = raw_value.strip().lower()
        if normalized in ("1", "true", "yes", "y", "on"):
            desired_status = True
        elif normalized in ("0", "false", "no", "n", "off"):
            desired_status = False
        else:
            raise HTTPException(status_code=400, detail="Некорректное значение is_verified")
    else:
        raise HTTPException(status_code=400, detail="Некорректное значение is_verified")

    if not await set_report_verification_status(report_id, desired_status):
        raise HTTPException(status_code=404, detail="Отчет не найден")

    return {
        "success": True,
        "data": {
            "id": report_id,
            "is_verified": desired_status
        }
    }

@app.delete("/api/report/{report_id}")
async def delete_report(report_id: int):
    """Удаляет отчет."""
    # Проверяем, существует ли отчет
    existing_report = await get_report_by_id(report_id)
    if existing_report is None:
        raise HTTPException(status_code=404, detail="Отчет не найден")
    
    success, message = await delete_report_from_db(report_id)
    if success:
        return {"success": True, "message": message}
    else:
        raise HTTPException(status_code=400, detail=message)


# Накопительная ведомость
@app.get("/api/accumulative-statement")
async def get_accumulative_statement(foreman_id: Optional[int] = None):
    """Получает накопительную ведомость выполненных работ."""
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            params = []
            foreman_filter = ""

            if foreman_id is not None:
                foreman_filter = " AND wr.foreman_id = ?"
                params.append(foreman_id)
            async with db.execute('''
                SELECT 
                    w.category AS Раздел,
                    w.name AS Работа,
                    w.unit AS Единица_измерения,
                    w.unit_cost_without_vat AS Стоимость_за_единицу,
                    SUM(wr.quantity) AS Количество,
                    w.project_total AS Проект,
                    CASE 
                        WHEN w.project_total > 0 THEN ROUND((SUM(wr.quantity) / w.project_total) * 100, 2)
                        ELSE 0
                    END AS Процент_выполнения,
                    SUM(wr.quantity * COALESCE(w.unit_cost_without_vat, 0)) AS Сумма
                FROM work_reports wr
                JOIN works w ON wr.work_id = w.id
                WHERE wr.is_verified = 1''' + foreman_filter + '''
                GROUP BY w.category, w.name, w.unit, w.project_total, w.unit_cost_without_vat
                ORDER BY w.category, w.name
            ''', params) as cursor:
                rows = await cursor.fetchall()
                accumulative_data = []
                for row in rows:
                    (
                        category,
                        work,
                        unit,
                        unit_cost_without_vat,
                        quantity,
                        project_total,
                        percentage,
                        total_without_vat,
                    ) = row
                    unit_cost_without_vat = unit_cost_without_vat or 0
                    total_without_vat = total_without_vat or 0
                    total_with_vat = round(total_without_vat * 1.2, 2)
                    accumulative_data.append({
                        'Раздел': category,
                        'Работа': work,
                        'Единица измерения': unit,
                        'Стоимость за единицу': unit_cost_without_vat,
                        'Количество': quantity,
                        'Проект': project_total,
                        '%Выполнения': percentage,
                        'Сумма (без НДС)': round(total_without_vat, 2),
                        'Сумма (с НДС)': total_with_vat,
                        'Сумма': round(total_without_vat, 2),
                    })
            async with db.execute('''
                SELECT DISTINCT f.id, f.first_name, f.last_name
                FROM work_reports wr
                JOIN foremen f ON wr.foreman_id = f.id
                WHERE wr.is_verified = 1
                ORDER BY f.first_name, f.last_name
            ''') as cursor:
                foremen_rows = await cursor.fetchall()
                available_foremen = [
                    {
                        'id': row[0],
                        'full_name': " ".join(part for part in [row[1], row[2]] if part).strip() or str(row[0])
                    }
                    for row in foremen_rows
                ]

            logger.info(f"📦 Загружена накопительная ведомость: {len(accumulative_data)} записей")
            return {"success": True, "data": accumulative_data, "foremen": available_foremen}
    except Exception as e:
        logger.error(f"❌ Ошибка получения накопительной ведомости: {e}")
        return {"success": False, "error": str(e)}

# Новые эндпоинты для работы с отчетами (work-reports)
@app.get("/api/work-reports")
async def get_work_reports():
    """Получает все отчеты о работах."""
    reports = await get_all_work_reports_from_db()
    return {"success": True, "data": reports}

@app.post("/api/work-reports")
async def create_work_report(request: Request):
    """Создает новый отчет о работе."""
    try:
        report_data = await request.json()
        
        # Поддержка отправки нескольких работ за один раз
        if isinstance(report_data, dict) and isinstance(report_data.get('works'), list):
            base_required = ["foreman_id", "report_date", "report_time"]
            for field in base_required:
                if field not in report_data:
                    raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")

            works_list = [item for item in report_data.get('works', []) if item is not None]
            if not works_list:
                raise HTTPException(status_code=400, detail="Добавьте хотя бы одну работу в отчет")

            created_ids: List[int] = []
            for work_item in works_list:
                if not isinstance(work_item, dict):
                    raise HTTPException(status_code=400, detail="Неверный формат данных работы")

                for field in ["work_id", "quantity"]:
                    if field not in work_item:
                        raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")

                work_id = work_item["work_id"]
                quantity = work_item["quantity"]

                if not isinstance(work_id, int) or work_id <= 0:
                    raise HTTPException(status_code=400, detail="work_id должно быть положительным целым числом")

                try:
                    quantity_value = float(quantity)
                except (TypeError, ValueError):
                    raise HTTPException(status_code=400, detail="quantity должно быть числом")

                if quantity_value <= 0:
                    raise HTTPException(status_code=400, detail="quantity должно быть больше 0")

                payload = {
                    "foreman_id": report_data["foreman_id"],
                    "work_id": work_id,
                    "quantity": quantity_value,
                    "report_date": report_data["report_date"],
                    "report_time": report_data["report_time"],
                }

                if "photo_report_url" in report_data:
                    payload["photo_report_url"] = report_data["photo_report_url"]

                success, result = await create_work_report_in_db(payload)
                if not success:
                    raise HTTPException(status_code=400, detail=result)

                created_ids.append(result)

            return {
                "success": True,
                "message": "Отчет успешно создан",
                "data": {"ids": created_ids},
            }

        # Обработка одиночного отчета (для обратной совместимости)
        required_fields = ["foreman_id", "work_id", "quantity", "report_date", "report_time"]
        for field in required_fields:
            if field not in report_data:
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")
        
        try:
            quantity_value = float(report_data["quantity"])
        except (TypeError, ValueError):
            raise HTTPException(status_code=400, detail="quantity должно быть числом")

        if quantity_value <= 0:
            raise HTTPException(status_code=400, detail="quantity должно быть больше 0")

        report_data["quantity"] = quantity_value


        success, result = await create_work_report_in_db(report_data)
        if success:
            return {"success": True, "message": "Отчет успешно создан", "data": {"id": result}}
        else:
            raise HTTPException(status_code=400, detail=result)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")
    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        logger.error(f"❌ Ошибка при создании отчета: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

@app.put("/api/work-reports/{report_id}")
async def update_work_report(report_id: int, request: Request):
    """Обновляет отчет о работе."""
    try:
        report_data = await request.json()
        
        # Валидация
        required_fields = ["foreman_id", "work_id", "quantity", "report_date", "report_time"]
        for field in required_fields:
            if field not in report_data:
                raise HTTPException(status_code=400, detail=f"Отсутствует поле: {field}")
        
        success, message = await update_work_report_in_db(report_id, report_data)
        if success:
            return {"success": True, "message": message}
        else:
            raise HTTPException(status_code=400, detail=message)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Неверный формат JSON")
    except HTTPException as http_exc:
        raise http_exc
    except Exception as e:
        logger.error(f"❌ Ошибка при обновлении отчета: {e}")
        raise HTTPException(status_code=500, detail="Внутренняя ошибка сервера")

# ========== ЗАПУСК СЕРВЕРА ==========
if __name__ == "__main__":
    import uvicorn
    logger.info(f"🚀 Запуск API сервера на {API_HOST}:{API_PORT}")
    uvicorn.run(app, host=API_HOST, port=API_PORT, log_level="info")